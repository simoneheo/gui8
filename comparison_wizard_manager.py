from PySide6.QtWidgets import QMainWindow, QMessageBox, QCheckBox, QLabel, QTableWidgetItem, QWidget, QHBoxLayout, QPushButton, QDialog
from PySide6.QtCore import Qt, QPoint
from PySide6.QtCore import Signal, QObject, QTimer
from PySide6.QtGui import QPainter, QColor, QPen, QBrush, QPixmap
import traceback
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas

from comparison_wizard_window import ComparisonWizardWindow
from channel import SourceType
from comparison import ComparisonRegistry, load_all_comparisons
from pair_analyzer import PairAnalyzer
from pair_analyzer import MethodConfigOp as PairAnalyzerMethodConfig
from overlay import Overlay
from overlay_wizard import OverlayWizard


class PairSelectionOp:
    """
    Operation class for intelligent channel pair selection
    Handles finding the best channel pairs for comparison
    """
    
    def __init__(self, file_manager, channel_manager, selected_file_id=None):
        self.file_manager = file_manager
        self.channel_manager = channel_manager
        self.selected_file_id = selected_file_id
        
    def find_intelligent_channel_pairs(self):
        """Find the best channel pairs using intelligent matching"""
        try:
            if not self.file_manager or not self.channel_manager:
                return None, None
                
            # PRIORITY 0: Use selected file from main window if available
            if self.selected_file_id:
                selected_file = self.file_manager.get_file(self.selected_file_id)
                if selected_file:
                    print(f"[PairSelectionOp] Prioritizing selected file: {selected_file.filename}")
                    return self._find_pairs_with_selected_file(selected_file)
            
            # Get all RAW channels from all files
            all_raw_channels = []
            for file_info in self.file_manager.get_all_files():
                file_channels = self.channel_manager.get_channels_by_file(file_info.file_id)
                raw_channels = [ch for ch in file_channels if ch.type == SourceType.RAW]
                all_raw_channels.extend(raw_channels)
            
            if len(all_raw_channels) < 2:
                return self._fallback_to_any_channels()
            
            # PRIORITY 1: Find channels with similar x_stats (min_val, max_val, count)
            similar_pairs = self._find_channels_with_similar_x_stats(all_raw_channels)
            if similar_pairs:
                print(f"[PairSelectionOp] Found {len(similar_pairs)} similar channel pairs")
                return similar_pairs[0][0], similar_pairs[0][1]  # Return first best match
            
            # PRIORITY 2: Find different RAW channels (any type)
            different_raw_pairs = self._find_different_raw_channels(all_raw_channels)
            if different_raw_pairs:
                print(f"[PairSelectionOp] Found {len(different_raw_pairs)} different RAW channel pairs")
                return different_raw_pairs[0]
            
            # PRIORITY 3: Fallback to any different channels
            return self._fallback_to_any_channels()
            
        except Exception as e:
            print(f"[PairSelectionOp] Error in intelligent channel selection: {e}")
            return None, None

    def _find_pairs_with_selected_file(self, selected_file):
        """Find channel pairs prioritizing the selected file"""
        try:
            # Get channels from selected file
            selected_file_channels = self.channel_manager.get_channels_by_file(selected_file.file_id)
            selected_raw_channels = [ch for ch in selected_file_channels if ch.type == SourceType.RAW]
            
            if not selected_raw_channels:
                return None, None
            
            # Get channels from other files
            other_file_channels = []
            for file_info in self.file_manager.get_all_files():
                if file_info.file_id != selected_file.file_id:
                    file_channels = self.channel_manager.get_channels_by_file(file_info.file_id)
                    raw_channels = [ch for ch in file_channels if ch.type == SourceType.RAW]
                    other_file_channels.extend(raw_channels)
            
            # PRIORITY: Find pairs between selected file and other files
            for selected_ch in selected_raw_channels:
                for other_ch in other_file_channels:
                    # Check if they have similar x_stats
                    stats1 = getattr(selected_ch, 'x_stats', None)
                    stats2 = getattr(other_ch, 'x_stats', None)
                    
                    if stats1 and stats2:
                        similarity = self._calculate_x_stats_similarity(stats1, stats2)
                        if similarity > 0.7:  # 70% similarity threshold for selected file
                            print(f"[PairSelectionOp] Found similar pair with selected file: {selected_ch.channel_id} vs {other_ch.channel_id}")
                            return selected_ch, other_ch
            
            # If no similar pairs found, use first channel from selected file with any other channel
            if other_file_channels:
                print(f"[PairSelectionOp] Using selected file channel with any other channel")
                return selected_raw_channels[0], other_file_channels[0]
            
            # If no other files, use two different channels from selected file
            if len(selected_raw_channels) >= 2:
                print(f"[PairSelectionOp] Using two channels from selected file")
                return selected_raw_channels[0], selected_raw_channels[1]
            
            return None, None
            
        except Exception as e:
            print(f"[PairSelectionOp] Error finding pairs with selected file: {e}")
            return None, None

    def _find_channels_with_similar_x_stats(self, channels):
        """Find channels with similar x_stats (min_val, max_val, count)"""
        similar_pairs = []
        
        for i, ch1 in enumerate(channels):
            stats1 = getattr(ch1, 'x_stats', None)
            if not stats1:
                continue
                
            for ch2 in channels[i+1:]:
                stats2 = getattr(ch2, 'x_stats', None)
                if not stats2:
                    continue
                
                # Calculate similarity score
                similarity = self._calculate_x_stats_similarity(stats1, stats2)
                
                if similarity > 0.8:  # 80% similarity threshold
                    similar_pairs.append((ch1, ch2, similarity))
        
        # Sort by similarity score (highest first)
        similar_pairs.sort(key=lambda x: x[2], reverse=True)
        return similar_pairs

    def _calculate_x_stats_similarity(self, stats1, stats2):
        """Calculate similarity between two x_stats objects"""
        try:
            # Normalize differences by range
            range1 = stats1.max_val - stats1.min_val
            range2 = stats2.max_val - stats2.min_val
            
            # Count similarity (exact match gets 1.0)
            count_similarity = 1.0 if stats1.count == stats2.count else 0.0
            
            # Range similarity (how close the ranges are)
            if range1 > 0 and range2 > 0:
                range_diff = abs(range1 - range2) / max(range1, range2)
                range_similarity = max(0, 1 - range_diff)
            else:
                range_similarity = 0.0
            
            # Min/Max similarity
            min_similarity = 1.0 if stats1.min_val == stats2.min_val else 0.0
            max_similarity = 1.0 if stats1.max_val == stats2.max_val else 0.0
            
            # Weighted average (count is most important)
            total_similarity = (count_similarity * 0.5 + 
                               range_similarity * 0.3 + 
                               min_similarity * 0.1 + 
                               max_similarity * 0.1)
            
            return total_similarity
        except Exception as e:
            print(f"Error calculating similarity: {e}")
            return 0.0

    def _find_different_raw_channels(self, channels):
        """Find different RAW channels from different files"""
        different_pairs = []
        
        for i, ch1 in enumerate(channels):
            for ch2 in channels[i+1:]:
                # Must be different channels from different files
                if (ch1.channel_id != ch2.channel_id and 
                    ch1.file_id != ch2.file_id):
                    different_pairs.append((ch1, ch2))
        
        return different_pairs

    def _fallback_to_any_channels(self):
        """Fallback to any available channels"""
        try:
            if not self.file_manager or not self.channel_manager:
                return None, None
                
            all_channels = []
            for file_info in self.file_manager.get_all_files():
                file_channels = self.channel_manager.get_channels_by_file(file_info.file_id)
                all_channels.extend(file_channels)
            
            if len(all_channels) >= 2:
                # Find different channels
                for i, ch1 in enumerate(all_channels):
                    for ch2 in all_channels[i+1:]:
                        if ch1.channel_id != ch2.channel_id:
                            return ch1, ch2
                
                # If no different channels found, use first two
                return all_channels[0], all_channels[1]
            elif len(all_channels) == 1:
                return all_channels[0], all_channels[0]
            else:
                return None, None
                
        except Exception as e:
            print(f"[PairSelectionOp] Error in fallback channel selection: {e}")
            return None, None


class DataAlignmentOp:
    """
    Operation class for intelligent data alignment parameter selection.
    Suggests alignment parameters based on selected channels.
    """
    def __init__(self, ref_channel, test_channel):
        self.ref_channel = ref_channel
        self.test_channel = test_channel

    def suggest_alignment(self, mode=None):
        """Main method to suggest alignment parameters based on mode"""
        if mode == 'Index-Based':
            return self.suggest_index_params()
        elif mode == 'Time-Based':
            return self.suggest_time_params()
        else:
            # Auto-detect best mode
            return self._auto_detect_mode()

    def suggest_index_params(self):
        """Suggest parameters for index-based alignment"""
        try:
            print(f"[DEBUG] suggest_index_params: ref_channel type={type(self.ref_channel)}")
            print(f"[DEBUG] suggest_index_params: test_channel type={type(self.test_channel)}")
            
            # Check if channels have ydata property
            ref_ydata = getattr(self.ref_channel, 'ydata', None)
            test_ydata = getattr(self.test_channel, 'ydata', None)
            
            print(f"[DEBUG] suggest_index_params: ref_ydata={ref_ydata}")
            print(f"[DEBUG] suggest_index_params: test_ydata={test_ydata}")
            
            ref_len = len(ref_ydata) if ref_ydata is not None else 0
            test_len = len(test_ydata) if test_ydata is not None else 0
            end_index = min(ref_len, test_len) - 1
            print(f"[DEBUG] suggest_index_params: ref_len={ref_len}, test_len={test_len}, end_index={end_index}")
            
            # Check if end_index is negative (which would cause issues)
            if end_index < 0:
                print(f"[DEBUG] suggest_index_params: end_index is negative ({end_index}), using 0")
                end_index = 0
                
            return {
                'mode': 'Index-Based',
                'start_index': 0,
                'end_index': end_index,
                'offset': 0,
            }
        except Exception as e:
            print(f"Error in suggest_index_params: {e}")
            return {
                'mode': 'Index-Based',
                'start_index': 0,
                'end_index': 100,
                'offset': 0,
            }

    def suggest_time_params(self):
        """Suggest parameters for time-based alignment"""
        try:
            # Validate that both channels have xrange
            if not (hasattr(self.ref_channel, 'xrange') and hasattr(self.test_channel, 'xrange') 
                   and self.ref_channel.xrange and self.test_channel.xrange):
                print("Warning: One or both channels missing xrange for time-based alignment")
                return self._default_time_params()
            
            ref_start, ref_end = self.ref_channel.xrange
            test_start, test_end = self.test_channel.xrange
            
            # Calculate overlap region
            overlap_start = max(ref_start, test_start)
            overlap_end = min(ref_end, test_end)
            
            if overlap_start >= overlap_end:
                print("Warning: No time overlap between channels")
                return self._default_time_params()
            
            return {
                'mode': 'Time-Based',
                'start_time': overlap_start,
                'end_time': overlap_end,
                'interpolation': self._suggest_interpolation(),
                'offset': 0,  # Always 0 unless user specifies otherwise
                'resolution': self._suggest_resolution(),  # Use intelligent resolution suggestion
            }
        except Exception as e:
            print(f"Error in suggest_time_params: {e}")
            return self._default_time_params()

    def _auto_detect_mode(self):
        """Auto-detect the best alignment mode based on channel properties"""
        # Try time-based first if both channels have xrange
        time_params = self.suggest_time_params()
        if time_params['mode'] == 'Time-Based':
            return time_params
        
        # Fallback to index-based
        return self.suggest_index_params()

    def _default_index_params(self):
        """Default parameters for index-based alignment"""
        try:
            ref_len = len(self.ref_channel.ydata) if self.ref_channel.ydata is not None else 0
            test_len = len(self.test_channel.ydata) if self.test_channel.ydata is not None else 0
            return {
                'mode': 'Index-Based',
                'start_index': 0,
                'end_index': min(ref_len, test_len) - 1,
                'offset': 0,
            }
        except Exception as e:
            print(f"Error in _default_index_params: {e}")
            return {
                'mode': 'Index-Based',
                'start_index': 0,
                'end_index': 100,
                'offset': 0,
            }

    def _default_time_params(self):
        """Default parameters for time-based alignment when channels don't have proper time info"""
        return {
            'mode': 'Time-Based',
            'start_time': 0.0,
            'end_time': 10.0,
            'interpolation': 'nearest',
            'offset': 0,  # Always 0 unless user specifies otherwise
            'resolution': self._suggest_resolution(),  # Use intelligent resolution suggestion
        }

    def _suggest_interpolation(self):
        """Suggest interpolation method based on channel properties"""
        try:
            sr1 = getattr(self.ref_channel, 'sampling_rate', None)
            sr2 = getattr(self.test_channel, 'sampling_rate', None)
            
            # Use linear interpolation if sampling rates are different
            if sr1 and sr2 and abs(sr1 - sr2) > 0.1:  # Allow small differences
                return 'linear'
            return 'nearest'
        except Exception as e:
            print(f"Error in _suggest_interpolation: {e}")
            return 'nearest'

    def _suggest_resolution(self):
        """Suggest resolution for time-based alignment based on sampling rates"""
        try:
            # Get sampling rates from both channels
            ref_sr = getattr(self.ref_channel, 'sampling_rate', None)
            test_sr = getattr(self.test_channel, 'sampling_rate', None)
            
            # If both channels have sampling rates, use the higher one
            if ref_sr is not None and test_sr is not None:
                # Use the higher sampling rate to preserve more detail
                max_sr = max(ref_sr, test_sr)
                # Resolution should be at least 2x the Nyquist frequency (1/2 of sampling rate)
                # but not too fine to avoid excessive data points
                suggested_resolution = 1.0 / max_sr
                
                # Clamp to reasonable bounds: between 0.001 and 1.0 seconds
                suggested_resolution = max(0.001, min(1.0, suggested_resolution))
                
                # Round to reasonable precision
                if suggested_resolution >= 0.1:
                    suggested_resolution = round(suggested_resolution, 1)  # 0.1, 0.2, 0.5, 1.0
                elif suggested_resolution >= 0.01:
                    suggested_resolution = round(suggested_resolution, 2)  # 0.01, 0.02, 0.05
                else:
                    suggested_resolution = round(suggested_resolution, 3)  # 0.001, 0.002, etc.
                
                return suggested_resolution
            
            # If only one channel has sampling rate, use that
            elif ref_sr is not None:
                suggested_resolution = max(0.001, min(1.0, 1.0 / ref_sr))
                return round(suggested_resolution, 3)
            elif test_sr is not None:
                suggested_resolution = max(0.001, min(1.0, 1.0 / test_sr))
                return round(suggested_resolution, 3)
            
            # If no sampling rate info, try to estimate from time data
            else:
                ref_x = getattr(self.ref_channel, 'xdata', None)
                test_x = getattr(self.test_channel, 'xdata', None)
                
                if ref_x is not None and len(ref_x) > 1:
                    # Estimate sampling rate from time differences
                    time_diffs = np.diff(ref_x)
                    if len(time_diffs) > 0:
                        avg_interval = np.mean(time_diffs)
                        if avg_interval > 0:
                            estimated_sr = 1.0 / avg_interval
                            suggested_resolution = max(0.001, min(1.0, 1.0 / estimated_sr))
                            return round(suggested_resolution, 3)
                
                if test_x is not None and len(test_x) > 1:
                    # Try test channel if ref channel didn't work
                    time_diffs = np.diff(test_x)
                    if len(time_diffs) > 0:
                        avg_interval = np.mean(time_diffs)
                        if avg_interval > 0:
                            estimated_sr = 1.0 / avg_interval
                            suggested_resolution = max(0.001, min(1.0, 1.0 / estimated_sr))
                            return round(suggested_resolution, 3)
            
            # Default fallback
            return 0.1
            
        except Exception as e:
            print(f"Error in _suggest_resolution: {e}")
            return 0.1


class MethodConfigOp:
    """
    Method Configuration Operation - captures current method configuration
    from comparison wizard to pass to PairAnalyzer.
    
    Simple data container that captures the current analysis settings.
    """
    
    def __init__(self, method_name: str, parameters: Dict[str, Any], 
                 plot_script: Optional[str], stats_script: Optional[str],
                 performance_options: Optional[Dict[str, Any]] = None):
        """
        Initialize MethodConfigOp with captured configuration.
        
        Args:
            method_name: Current selected method name
            parameters: Current parameter values from wizard
            plot_script: Current plot script source code (None if not modified)
            stats_script: Current stats script source code (None if not modified)
            performance_options: Performance options from wizard (max_points, density_mode, etc.)
        """
        self.method_name = method_name
        self.parameters = parameters
        self.plot_script = plot_script
        self.stats_script = stats_script
        self.performance_options = performance_options or {}
        self.timestamp = datetime.now()
    
    @classmethod
    def from_wizard(cls, comparison_wizard) -> 'MethodConfigOp':
        """
        Create MethodConfigOp by capturing current state from comparison wizard.
        Only captures script content if it has been modified from the original.
        
        Args:
            comparison_wizard: ComparisonWizardWindow instance
            
        Returns:
            MethodConfigOp with current wizard configuration
        """
        try:
            # Capture method name
            method_name = comparison_wizard.get_current_method_name() or "Unknown"
            
            # Capture parameters
            parameters = comparison_wizard.get_current_parameters()
            
            # Capture scripts only if modified
            current_plot_script = comparison_wizard.plot_script_text.toPlainText() or ""
            current_stats_script = comparison_wizard.stats_script_text.toPlainText() or ""
            
            # Use ScriptChangeTracker to check if scripts have been modified
            plot_script = None
            stats_script = None
            
            if comparison_wizard.script_tracker.is_plot_script_modified(current_plot_script):
                plot_script = current_plot_script
                
            if comparison_wizard.script_tracker.is_stats_script_modified(current_stats_script):
                stats_script = current_stats_script
            
            return cls(
                method_name=method_name,
                parameters=parameters,
                plot_script=plot_script,
                stats_script=stats_script
            )
            
        except Exception as e:
            print(f"[MethodConfigOp] Error capturing wizard config: {e}")
            # Return default config on error
            return cls(
                method_name="Unknown",
                parameters={},
                plot_script=None,
                stats_script=None
            )
    
    def to_dict(self) -> Dict[str, Any]:
        """
        Convert MethodConfigOp to dictionary for debugging or serialization.
        
        Returns:
            Dictionary representation of the method configuration
        """
        return {
            'method_name': self.method_name,
            'parameters': self.parameters,
            'plot_script_length': len(self.plot_script) if self.plot_script else 0,
            'stats_script_length': len(self.stats_script) if self.stats_script else 0,
            'timestamp': self.timestamp.isoformat()
        }
    
    def get_plot_script(self) -> str:
        """
        Get the plot script to use (modified or default).
        
        Returns:
            Modified script if available, otherwise default script from comparison class
        """
        if self.plot_script:
            return self.plot_script  # Use modified script
        else:
            # Get default script from comparison class
            try:
                comparison_cls = ComparisonRegistry.get(self.method_name)
                if comparison_cls and hasattr(comparison_cls, 'plot_script'):
                    import inspect
                    return inspect.getsource(comparison_cls.plot_script)
                else:
                    return "# No default plot script available"
            except Exception as e:
                print(f"[MethodConfigOp] Error getting default plot script: {e}")
                return "# Error loading default plot script"
    
    def get_stats_script(self) -> str:
        """
        Get the stats script to use (modified or default).
        
        Returns:
            Modified script if available, otherwise default script from comparison class
        """
        if self.stats_script:
            return self.stats_script  # Use modified script
        else:
            # Get default script from comparison class
            try:
                comparison_cls = ComparisonRegistry.get(self.method_name)
                if comparison_cls and hasattr(comparison_cls, 'stats_script'):
                    import inspect
                    return inspect.getsource(comparison_cls.stats_script)
                else:
                    return "# No default stats script available"
            except Exception as e:
                print(f"[MethodConfigOp] Error getting default stats script: {e}")
                return "# Error loading default stats script"
    
    def has_modified_plot_script(self) -> bool:
        """Check if plot script has been modified from default"""
        return self.plot_script is not None
    
    def has_modified_stats_script(self) -> bool:
        """Check if stats script has been modified from default"""
        return self.stats_script is not None
    
    def __str__(self) -> str:
        """String representation for debugging"""
        return f"MethodConfigOp(method={self.method_name}, params={len(self.parameters)}, timestamp={self.timestamp})"


class RenderPlotOp:
    """
    Handles plot rendering for comparison analysis results.
    Separates computation (PairAnalyzer) from visualization (matplotlib).
    """
    
    def __init__(self, plot_widget=None):
        """
        Initialize the render plot operation.
        
        Args:
            plot_widget: Widget containing the matplotlib figure/axes
        """
        self.plot_widget = plot_widget
        self.current_figure = None
        self.current_axes = None
        
        # Store artist references for efficient visibility toggling
        self.pair_artists = {}  # pair_id -> matplotlib artist
        self.overlay_artists = {}  # overlay_id -> matplotlib artist
        
        # Store scatter data for overlay calculations
        self.current_scatter_data = []  # List of scatter data dictionaries
        
        # Initialize overlay zorder to ensure overlays appear above scatter points
        self._current_overlay_zorder = 10
    
    def render(self, analysis_results: Dict[str, Any], plot_config: Dict[str, Any] = None) -> bool:
        """
        Render analysis results to the plot area.
        
        Args:
            analysis_results: Results from PairAnalyzer containing scatter_data and overlays
            plot_config: Configuration for plot appearance and behavior
            
        Returns:
            bool: True if rendering successful, False otherwise
        """
        try:
            if not analysis_results:
                print("[RenderPlotOp] No analysis results to render")
                return False
                
            if plot_config is None:
                plot_config = {}
                
            # Get data from analysis results
            scatter_data = analysis_results.get('scatter_data', [])
            overlays = analysis_results.get('overlays', [])
            method_name = analysis_results.get('method_name', 'unknown')
            plot_type = analysis_results.get('plot_type', 'scatter')
            
            print(f"[RenderPlotOp] Rendering with plot_type: {plot_type}")
            
            # Apply performance options to scatter data BEFORE rendering (only for scatter plots)
            if plot_config and plot_type == "scatter":
                scatter_data = self._apply_performance_options(scatter_data, plot_config)
            
            # Store scatter data for overlay calculations (make a copy to avoid reference issues)
            self.current_scatter_data = scatter_data.copy()
            
            print(f"[RenderPlotOp] Rendering {len(scatter_data)} pairs with {len(overlays)} overlays")
            
            # Clear previous plot
            self._clear_plot()
            
            # Create new figure and axes if needed
            if not self._ensure_plot_ready():
                return False
                
            # Route rendering based on plot_type
            try:
                if plot_type == "scatter":
                    self._render_scatter_plot(scatter_data, plot_config)
                elif plot_type == "bar":
                    self._render_bar_plot(scatter_data, plot_config)
                elif plot_type == "stacked_area":
                    self._render_stacked_area_plot(scatter_data, plot_config)
                elif plot_type == "histogram":
                    self._render_histogram_plot(scatter_data, plot_config)
                elif plot_type == "line":
                    self._render_line_plot(scatter_data, plot_config)
                else:
                    print(f"[RenderPlotOp] Unknown plot_type: {plot_type}, falling back to scatter")
                    self._render_scatter_plot(scatter_data, plot_config)
            except Exception as e:
                print(f"[RenderPlotOp] Error rendering main plot: {e}")
                import traceback
                traceback.print_exc()
                # Continue with overlays even if main plot fails
            
            # Auto-zoom based on plot type
            try:
                if plot_type == "histogram":
                    self._auto_zoom_to_histogram_data(scatter_data)
                else:
                    self._auto_zoom_to_scatter_data(scatter_data)
            except Exception as e:
                print(f"[RenderPlotOp] Error auto-zooming: {e}")
                # Continue without auto-zoom
            
            # Render overlays
            print(f"[RenderPlotOp] About to render {len(overlays)} overlays")
            for overlay in overlays:
                print(f"[RenderPlotOp] Overlay {overlay.id}: type={overlay.type}, data={overlay.data}, show={overlay.show}")
            self._render_overlays(overlays, plot_config)
            
            # Configure plot appearance
            try:
                self._configure_plot_appearance(scatter_data, method_name, plot_config)
            except Exception as e:
                print(f"[RenderPlotOp] Error configuring plot appearance: {e}")
                # Continue without appearance configuration
            
            # Refresh the plot widget
            try:
                self._refresh_plot_widget()
            except Exception as e:
                print(f"[RenderPlotOp] Error refreshing plot widget: {e}")
                # Try alternative refresh method
                try:
                    if self.current_figure and hasattr(self.current_figure, 'canvas'):
                        self.current_figure.canvas.draw()
                        print(f"[RenderPlotOp] Used alternative refresh method")
                except Exception as e2:
                    print(f"[RenderPlotOp] Alternative refresh also failed: {e2}")
            
            print(f"[RenderPlotOp] Successfully rendered plot for method: {method_name}")
            return True
            
        except Exception as e:
            print(f"[RenderPlotOp] Critical error rendering plot: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _clear_plot(self):
        """Clear the current plot."""
        if self.current_axes:
            self.current_axes.clear()
        
        # Clear artist references
        self.pair_artists.clear()
        self.overlay_artists.clear()
        
        # Clear stored scatter data
        self.current_scatter_data.clear()
    
    def _ensure_plot_ready(self) -> bool:
        """Ensure plot figure and axes are ready for rendering."""
        try:
            if self.plot_widget and hasattr(self.plot_widget, 'figure'):
                self.current_figure = self.plot_widget.figure
                # Clear existing axes and create new one
                self.current_figure.clear()
                self.current_axes = self.current_figure.add_subplot(111)
                return True
            else:
                # Fallback: create standalone figure
                self.current_figure = plt.figure(figsize=(10, 6))
                self.current_axes = self.current_figure.add_subplot(111)
                return True
        except Exception as e:
            print(f"[RenderPlotOp] Error setting up plot: {e}")
            return False
    
    def _render_scatter_plot(self, scatter_data: List[Dict[str, Any]], plot_config: Dict[str, Any]):
        """Render scatter points with performance options (density modes)."""
        if not scatter_data:
            return
            
        for i, pair_data in enumerate(scatter_data):
            try:
                x_data = pair_data.get('x_data', [])
                y_data = pair_data.get('y_data', [])
                
                if len(x_data) == 0 or len(y_data) == 0:
                    continue
                
                # Get density mode from pair data (set by performance options)
                density_mode = pair_data.get('density_mode', 'scatter')
                
                # Render based on density mode
                if density_mode == 'scatter':
                    self._render_scatter_points(pair_data)
                elif density_mode == 'hexbin':
                    self._render_hexbin_plot(pair_data)
                elif density_mode == 'kde':
                    self._render_kde_plot(pair_data)
                else:
                    print(f"[RenderPlotOp] Unknown density mode: {density_mode}, falling back to scatter")
                    self._render_scatter_points(pair_data)
                
            except Exception as e:
                print(f"[RenderPlotOp] ERROR: Error rendering pair {pair_data.get('pair_name', 'unknown')}: {e}")
                import traceback
                traceback.print_exc()
    
    def _auto_zoom_to_scatter_data(self, scatter_data: List[Dict[str, Any]], padding_factor: float = 0.05):
        """Auto-zoom the plot to focus on scatter data, excluding overlays."""
        if not scatter_data or not self.current_axes:
            return
        
        # Calculate bounds from all scatter data
        x_min, x_max = float('inf'), float('-inf')
        y_min, y_max = float('inf'), float('-inf')
        
        for pair_data in scatter_data:
            x_data = pair_data.get('x_data', [])
            y_data = pair_data.get('y_data', [])
            
            # Check if arrays have data (handle both lists and numpy arrays)
            if len(x_data) > 0 and len(y_data) > 0:
                # Handle numpy arrays or lists
                try:
                    x_min = min(x_min, min(x_data))
                    x_max = max(x_max, max(x_data))
                    y_min = min(y_min, min(y_data))
                    y_max = max(y_max, max(y_data))
                except (ValueError, TypeError):
                    # Skip if data is not numeric
                    continue
        
        # Check if we found valid bounds
        if x_min == float('inf') or x_max == float('-inf') or y_min == float('inf') or y_max == float('-inf'):
            print("[RenderPlotOp] No valid scatter data bounds found, using auto-limits")
            return
        
        # Handle edge case where all data points are identical
        x_range = x_max - x_min
        y_range = y_max - y_min
        
        if x_range == 0:
            x_range = abs(x_min) * 0.1 if x_min != 0 else 1.0
            x_min -= x_range / 2
            x_max += x_range / 2
        
        if y_range == 0:
            y_range = abs(y_min) * 0.1 if y_min != 0 else 1.0
            y_min -= y_range / 2
            y_max += y_range / 2
        
        # Apply padding
        x_padding = x_range * padding_factor
        y_padding = y_range * padding_factor
        
        # Set axis limits
        self.current_axes.set_xlim(x_min - x_padding, x_max + x_padding)
        self.current_axes.set_ylim(y_min - y_padding, y_max + y_padding)
        
        print(f"[RenderPlotOp] Auto-zoomed to scatter data bounds: x=({x_min:.3f}, {x_max:.3f}), y=({y_min:.3f}, {y_max:.3f})")
    
    def _auto_zoom_to_histogram_data(self, scatter_data: List[Dict[str, Any]], padding_factor: float = 0.05):
        """Auto-zoom the plot to histogram data with proper y-range for frequencies."""
        if not scatter_data or not self.current_axes:
            return
        
        print(f"[RenderPlotOp] Auto-zooming for histogram with {len(scatter_data)} data series")
        
        # Calculate x-range from all histogram data (error values)
        x_min, x_max = float('inf'), float('-inf')
        max_frequency = 0
        
        for pair_data in scatter_data:
            x_data = pair_data.get('x_data', [])
            
            if len(x_data) > 0:
                try:
                    # X-range: based on the error values being histogrammed
                    x_min = min(x_min, min(x_data))
                    x_max = max(x_max, max(x_data))
                    
                    # Estimate maximum frequency by calculating histogram
                    bins = pair_data.get('bins', 30)
                    bin_edges = pair_data.get('bin_edges', None)
                    
                    if bin_edges is not None:
                        # Use pre-computed bin edges
                        counts, _ = np.histogram(x_data, bins=bin_edges)
                    else:
                        # Use bin count
                        counts, _ = np.histogram(x_data, bins=bins)
                    
                    # Track maximum frequency across all pairs
                    max_frequency = max(max_frequency, max(counts) if len(counts) > 0 else 0)
                    
                except (ValueError, TypeError):
                    # Skip if data is not numeric
                    continue
        
        # Check if we found valid bounds
        if x_min == float('inf') or x_max == float('-inf'):
            print("[RenderPlotOp] No valid histogram data bounds found, using auto-limits")
            return
        
        # Handle edge case where all data points are identical
        x_range = x_max - x_min
        if x_range == 0:
            x_range = abs(x_min) * 0.1 if x_min != 0 else 1.0
            x_min -= x_range / 2
            x_max += x_range / 2
        
        # Apply padding to x-range
        x_padding = x_range * padding_factor
        
        # For histograms: y always starts at 0, goes to max_frequency + padding
        y_min = 0
        y_max = max_frequency * (1 + padding_factor * 2)  # Extra padding for readability
        
        # Ensure minimum y-range
        if y_max <= y_min:
            y_max = y_min + 1
        
        # Set axis limits
        self.current_axes.set_xlim(x_min - x_padding, x_max + x_padding)
        self.current_axes.set_ylim(y_min, y_max)
        
        print(f"[RenderPlotOp] Auto-zoomed histogram: x=({x_min:.3f}, {x_max:.3f}), y=({y_min:.3f}, {y_max:.3f})")
        print(f"[RenderPlotOp] Max frequency: {max_frequency}, Y-range: [0, {y_max:.3f}]")
    
    def _render_overlays(self, overlays: List[Overlay], plot_config: Dict[str, Any]):
        """Render overlay elements on the plot."""
        for overlay in overlays:
            try:
                # Render ALL overlays to create artist references
                artist = self._render_single_overlay(overlay, plot_config)
                
                # Set initial visibility based on overlay.show property
                if artist and hasattr(artist, 'set_visible'):
                    artist.set_visible(overlay.show)
                    print(f"[RenderPlotOp] Set initial visibility for {overlay.id}: {overlay.show}")
                elif artist and isinstance(artist, list):
                    # Handle multiple artists (e.g., multiple lines)
                    for a in artist:
                        if hasattr(a, 'set_visible'):
                            a.set_visible(overlay.show)
                    print(f"[RenderPlotOp] Set initial visibility for {overlay.id} (multiple artists): {overlay.show}")
                
            except Exception as e:
                print(f"[RenderPlotOp] Error rendering overlay {overlay.id}: {e}")
                import traceback
                traceback.print_exc()
                # Continue with next overlay instead of breaking the entire plot
                continue
    
    def _render_single_overlay(self, overlay: Overlay, plot_config: Dict[str, Any]):
        """Render a single overlay element with rich styling support."""
        style = overlay.style
        overlay_type = overlay.type
        overlay_data = overlay.data or {}
        
        print(f"[RenderPlotOp] Rendering overlay: {overlay.id} (type: {overlay_type})")
        print(f"[RenderPlotOp] Overlay data keys: {list(overlay_data.keys())}")
        print(f"[RenderPlotOp] Overlay style keys: {list(style.keys())}")
        
        # Store artist reference for this overlay
        artist = None
        
        if overlay_type == 'line':
            artist = self._render_line_overlay(overlay, style)
        elif overlay_type == 'hline':
            # Handle hline type overlays using the line renderer
            artist = self._render_line_overlay(overlay, style)
        elif overlay_type == 'vline':
            # Handle vline type overlays using the line renderer
            artist = self._render_line_overlay(overlay, style)
        elif overlay_type == 'text':
            # Try to render with detailed error handling
            artist = self._render_text_overlay(overlay, style)
        elif overlay_type == 'fill':
            artist = self._render_fill_overlay(overlay, style)
        elif overlay_type == 'legend':
            artist = self._render_legend_overlay(overlay, style)
        else:
            print(f"[RenderPlotOp] Unknown overlay type: {overlay_type}")
        
        # Store artist reference for visibility toggling
        if artist:
            self.overlay_artists[overlay.id] = artist
            print(f"[RenderPlotOp] Registered artist for overlay {overlay.id}")
        else:
            print(f"[RenderPlotOp] No artist created for overlay {overlay.id}")
        
        return artist
    
    def _render_line_overlay(self, overlay: Overlay, style: Dict[str, Any]):
        """Render line-type overlays (horizontal, vertical, or general lines)."""
        if not self.current_axes:
            print("[RenderPlotOp] No axes available for line overlay")
            return None
        color = style.get('color')
        linestyle = style.get('linestyle')
        linewidth = style.get('linewidth')
        alpha = style.get('alpha')
        label = style.get('label', overlay.name)
        zorder = style.get('zorder', getattr(self, '_current_overlay_zorder', 10))
        artists = []
        overlay_data = overlay.data or {}
        
        print(f"[RenderPlotOp] Line overlay data: {overlay_data}")
        print(f"[RenderPlotOp] Line overlay style: {style}")
        
        # Check overlay_data first for coordinate data
        if 'y_value' in overlay_data:
            y_value = overlay_data['y_value']
            artist = self.current_axes.axhline(
                y=y_value, color=color, linestyle=linestyle,
                linewidth=linewidth, alpha=alpha, label=label, zorder=zorder
            )
            artists.append(artist)
            print(f"[RenderPlotOp] Added horizontal line at y={y_value}")
        elif 'y' in overlay_data and (overlay.type == 'hline' or 'hline' in overlay.id.lower() or 'bias' in overlay.name.lower()):
            # Handle hline type overlays with {'y': constant_value} format
            y_value = overlay_data['y']
            artist = self.current_axes.axhline(
                y=y_value, color=color, linestyle=linestyle,
                linewidth=linewidth, alpha=alpha, label=label, zorder=zorder
            )
            artists.append(artist)
            print(f"[RenderPlotOp] Added horizontal line (hline type) at y={y_value}")
        elif 'x' in overlay_data and (overlay.type == 'vline' or 'vline' in overlay.id.lower() or 'mean' in overlay.name.lower() or 'median' in overlay.name.lower()):
            # Handle vline type overlays with {'x': [x1, x2, ...]} format
            x_values = overlay_data['x']
            if not isinstance(x_values, list):
                x_values = [x_values]
            
            for i, x_value in enumerate(x_values):
                line_label = label if i == 0 else None
                artist = self.current_axes.axvline(
                    x=x_value, color=color, linestyle=linestyle,
                    linewidth=linewidth, alpha=alpha, label=line_label, zorder=zorder
                )
                artists.append(artist)
            print(f"[RenderPlotOp] Added {len(x_values)} vertical lines (vline type) at x={x_values}")
        elif 'x_value' in overlay_data:
            x_value = overlay_data['x_value']
            artist = self.current_axes.axvline(
                x=x_value, color=color, linestyle=linestyle,
                linewidth=linewidth, alpha=alpha, label=label, zorder=zorder
            )
            artists.append(artist)
            print(f"[RenderPlotOp] Added vertical line at x={x_value}")
        elif 'y_values' in overlay_data:
            y_values = overlay_data['y_values']
            for i, y_value in enumerate(y_values):
                line_label = label if i == 0 else None
                artist = self.current_axes.axhline(
                    y=y_value, color=color, linestyle=linestyle,
                    linewidth=linewidth, alpha=alpha, label=line_label, zorder=zorder
                )
                artists.append(artist)
            print(f"[RenderPlotOp] Added {len(y_values)} horizontal lines")
        elif 'x_values' in overlay_data:
            x_values = overlay_data['x_values']
            for i, x_value in enumerate(x_values):
                line_label = label if i == 0 else None
                artist = self.current_axes.axvline(
                    x=x_value, color=color, linestyle=linestyle,
                    linewidth=linewidth, alpha=alpha, label=line_label, zorder=zorder
                )
                artists.append(artist)
            print(f"[RenderPlotOp] Added {len(x_values)} vertical lines")
        elif ('x' in overlay_data and 'y' in overlay_data) or ('x_data' in style and 'y_data' in style):
            if 'x' in overlay_data and 'y' in overlay_data:
                x_data = overlay_data['x']
                y_data = overlay_data['y']
            else:
                x_data = style['x_data']
                y_data = style['y_data']
            if label == 'Regression Line' and 'slope' in style and 'intercept' in style:
                x_min = float('inf')
                x_max = float('-inf')
                for pair_data in self.current_scatter_data:
                    pair_x_data = pair_data.get('x_data', [])
                    if pair_x_data:
                        x_min = min(x_min, min(pair_x_data))
                        x_max = max(x_max, max(pair_x_data))
                if x_min == float('inf') or x_max == float('-inf'):
                    x_min, x_max = -10, 10
                slope = style['slope']
                intercept = style['intercept']
                x_data = [x_min, x_max]
                y_data = [slope * x_min + intercept, slope * x_max + intercept]
                print(f"[RenderPlotOp] Computed regression line: y = {slope:.3f}x + {intercept:.3f} over range [{x_min:.3f}, {x_max:.3f}]")
            artist = self.current_axes.plot(
                x_data, y_data, color=color, linestyle=linestyle,
                linewidth=linewidth, alpha=alpha, label=label, zorder=zorder
            )
            artists.append(artist[0])
            print(f"[RenderPlotOp] Added general line with {len(x_data)} points")
        elif label == 'y = x' or 'identity' in overlay.id.lower():
            x_min, x_max = self.current_axes.get_xlim()
            artist = self.current_axes.plot(
                [x_min, x_max], [x_min, x_max], 
                color=color, linestyle=linestyle,
                linewidth=linewidth, alpha=alpha, label=label, zorder=zorder
            )
            artists.append(artist[0])
            print(f"[RenderPlotOp] Added identity line")
        if len(artists) == 0:
            return None
        elif len(artists) == 1:
            return artists[0]
        else:
            return artists
    
    def _render_text_overlay(self, overlay: Overlay, style: Dict[str, Any]):
        """Render text-type overlays (statistical results, equations, etc.)."""
        if not self.current_axes:
            print("[RenderPlotOp] No axes available for text overlay")
            return None
            
        try:
            # Get text content from overlay data
            overlay_data = overlay.data or {}
            
            # Try multiple ways to get text content
            text = None
            
            # Method 1: Check for text_lines list
            text_lines = overlay_data.get('text_lines', style.get('text_lines', []))
            if text_lines:
                text = '\n'.join(str(line) for line in text_lines if line)
            
            # Method 2: Check for raw text string
            if not text and 'text' in overlay_data:
                text = str(overlay_data['text'])
            
            # Method 3: Check for raw text in style
            if not text and 'text' in style:
                text = str(style['text'])
            
            # Method 4: Use overlay name as fallback
            if not text:
                text = overlay.name
            
            if not text.strip():
                print(f"[RenderPlotOp] Empty text content for overlay {overlay.id}")
                return None
            
            # Clean problematic characters that matplotlib might misinterpret
            cleaned_text = text
            cleaned_text = cleaned_text.replace('²', '^2')  # Replace superscript 2
            cleaned_text = cleaned_text.replace('°', 'deg')  # Replace degree symbol
            cleaned_text = cleaned_text.replace('±', '+/-')  # Replace plus-minus symbol
            
            # Use the overlay's built-in rendering method
            overlay._render_text(self.current_axes)
            
            # Get the last text artist that was added (most recent one)
            if self.current_axes.texts:
                artist = self.current_axes.texts[-1]  # Get the most recently added text artist
                print(f"[RenderPlotOp] Successfully created text artist for overlay {overlay.id}")
                return artist
            else:
                print(f"[RenderPlotOp] No text artist found after rendering overlay {overlay.id}")
                return None
                
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering text overlay '{overlay.id}': {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _render_fill_overlay(self, overlay: Overlay, style: Dict[str, Any]):
        """Render fill-type overlays (confidence intervals, bands, etc.)."""
        if not self.current_axes:
            print("[RenderPlotOp] No axes available for fill overlay")
            return None
        color = style.get('color')
        alpha = style.get('alpha')
        edgecolor = style.get('edgecolor')
        linewidth = style.get('linewidth')
        label = style.get('label', overlay.name)
        zorder = style.get('zorder', getattr(self, '_current_overlay_zorder', 10))
        artist = None
        overlay_data = overlay.data or {}
        try:
            if 'confidence_intervals' in style:
                ci_data = style['confidence_intervals']
                artist = self._render_confidence_intervals(ci_data, color or '#ff69b4', alpha or 0.1, label)
            elif ('y_lower' in overlay_data and 'y_upper' in overlay_data) or ('y_lower' in style and 'y_upper' in style):
                if 'y_lower' in overlay_data and 'y_upper' in overlay_data:
                    y_lower = overlay_data['y_lower']
                    y_upper = overlay_data['y_upper']
                    x_coords = overlay_data.get('x')
                else:
                    y_lower = style['y_lower']
                    y_upper = style['y_upper']
                    x_coords = style.get('x')
                if x_coords is not None and len(x_coords) == len(y_lower) == len(y_upper):
                    artist = self.current_axes.fill_between(
                        x_coords, y_lower, y_upper,
                        alpha=alpha, color=color, edgecolor=edgecolor,
                        linewidth=linewidth, label=label, zorder=zorder
                    )
                else:
                    x_min, x_max = self.current_axes.get_xlim()
                    artist = self.current_axes.fill_between(
                        [x_min, x_max], y_lower, y_upper,
                        alpha=alpha, color=color, edgecolor=edgecolor,
                        linewidth=linewidth, label=label, zorder=zorder
                    )
                print(f"[RenderPlotOp] Added confidence bands with {len(y_lower)} points")
            elif ('x_lower' in overlay_data and 'x_upper' in overlay_data) or ('x_lower' in style and 'x_upper' in style):
                if 'x_lower' in overlay_data and 'x_upper' in overlay_data:
                    x_lower = overlay_data['x_lower']
                    x_upper = overlay_data['x_upper']
                else:
                    x_lower = style['x_lower']
                    x_upper = style['x_upper']
                y_min, y_max = self.current_axes.get_ylim()
                artist = self.current_axes.fill_betweenx(
                    [y_min, y_max], x_lower, x_upper,
                    alpha=alpha, color=color, edgecolor=edgecolor,
                    linewidth=linewidth, label=label, zorder=zorder
                )
                print(f"[RenderPlotOp] Added fill between x={x_lower} and x={x_upper}")
            elif 'confidence_level' in style:
                confidence_level = style['confidence_level']
                print(f"[RenderPlotOp] Confidence bands not implemented for level {confidence_level}")
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering fill overlay '{overlay.id}': {e}")
            return None
        return artist
    
    def _render_confidence_intervals(self, ci_data: Dict[str, Any], color: str, alpha: float, label: str):
        """Render confidence intervals from the data structure."""
        x_min, x_max = self.current_axes.get_xlim()
        
        artists = []
        
        # Handle different CI data structures
        if isinstance(ci_data, dict):
            for ci_name, ci_bounds in ci_data.items():
                if isinstance(ci_bounds, (list, tuple)) and len(ci_bounds) == 2:
                    lower, upper = ci_bounds
                    try:
                        artist = self.current_axes.fill_between(
                            [x_min, x_max], lower, upper,
                            alpha=alpha, color=color, 
                            label=f"{label} ({ci_name})" if ci_name != 'default' else label
                        )
                        artists.append(artist)
                    except Exception as e:
                        print(f"[RenderPlotOp] Error adding CI for '{ci_name}': {e}")
                elif isinstance(ci_bounds, dict):
                    # Nested structure like {'bias_ci': (lower, upper)}
                    for sub_name, sub_bounds in ci_bounds.items():
                        if isinstance(sub_bounds, (list, tuple)) and len(sub_bounds) == 2:
                            lower, upper = sub_bounds
                            try:
                                artist = self.current_axes.fill_between(
                                    [x_min, x_max], lower, upper,
                                    alpha=alpha, color=color,
                                    label=f"{label} ({sub_name})"
                                )
                                artists.append(artist)
                            except Exception as e:
                                print(f"[RenderPlotOp] Error adding nested CI for '{sub_name}': {e}")
        
        print(f"[RenderPlotOp] Added {len(artists)} confidence intervals")
        
        # Return single artist or list of artists
        if len(artists) == 0:
            return None
        elif len(artists) == 1:
            return artists[0]
        else:
            return artists
    
    def _render_legend_overlay(self, overlay: Overlay, style: Dict[str, Any]):
        """Render legend overlay - disabled to never show legends."""
        # Never show legend (disabled as requested)
        # visible = style.get('visible', True)
        # location = style.get('location', 'best')
        # fontsize = style.get('fontsize', 9)
        
        artist = None
        
        # Remove any existing legend
        if self.current_axes.get_legend():
            self.current_axes.get_legend().remove()
            print(f"[RenderPlotOp] Removed legend (disabled)")
        
        return artist
    
    def _configure_plot_appearance(self, scatter_data: List[Dict[str, Any]], 
                                 method_name: str, plot_config: Dict[str, Any]):
        """Configure plot labels, title, grid, legend, etc."""
        if not self.current_axes:
            return
            
        # Set labels from first pair's metadata if available
        if scatter_data:
            metadata = scatter_data[0].get('metadata', {})
            x_label = metadata.get('x_label', 'X Values')
            y_label = metadata.get('y_label', 'Y Values')
            title = metadata.get('title', f'{method_name.title()}')
        else:
            x_label = 'X Values'
            y_label = 'Y Values'
            title = f'{method_name.title()}'
        
        self.current_axes.set_xlabel(x_label)
        self.current_axes.set_ylabel(y_label)
        self.current_axes.set_title(title)
        
        # Add grid
        self.current_axes.grid(True, alpha=0.3)
        
        # Never show legend (disabled as requested)
        # handles, labels = self.current_axes.get_legend_handles_labels()
        # if len(handles) > 1:
        #     self.current_axes.legend(loc='best', fontsize=8)
    
    def _refresh_plot_widget(self):
        """Refresh the plot widget to show changes."""
        try:
            if self.plot_widget and hasattr(self.plot_widget, 'draw'):
                self.plot_widget.draw()
                print(f"[RenderPlotOp] Refreshed canvas widget")
            elif self.current_figure:
                self.current_figure.tight_layout()
                if hasattr(self.current_figure, 'canvas'):
                    self.current_figure.canvas.draw()
                    print(f"[RenderPlotOp] Refreshed standalone figure")
        except Exception as e:
            print(f"[RenderPlotOp] Error refreshing plot widget: {e}")
    
    def toggle_pair_visibility(self, pair_id: str, visible: bool):
        """Toggle visibility of a pair's scatter plot without recomputing."""
        if pair_id not in self.pair_artists:
            print(f"[RenderPlotOp] No artist found for pair {pair_id}")
            return False
        
        try:
            artist = self.pair_artists[pair_id]
            artist.set_visible(visible)
            self._refresh_plot_widget()
            print(f"[RenderPlotOp] Toggled pair {pair_id} visibility to {visible}")
            return True
        except Exception as e:
            print(f"[RenderPlotOp] Error toggling pair visibility: {e}")
            return False
    
    def update_statistical_overlays(self, overlays: List[Overlay]):
        """Update only statistical overlays without recomputing scatter data."""
        if not self.current_axes:
            print("[RenderPlotOp] No axes available for overlay update")
            return False
        
        try:
            # Remove existing statistical overlays
            for overlay_id in list(self.overlay_artists.keys()):
                if overlay_id in self.overlay_artists:
                    artist = self.overlay_artists[overlay_id]
                    if hasattr(artist, 'remove'):
                        artist.remove()
                    elif isinstance(artist, list):
                        for a in artist:
                            if hasattr(a, 'remove'):
                                a.remove()
                    del self.overlay_artists[overlay_id]
            
            # Render ALL overlays to create artist references
            for overlay in overlays:
                artist = self._render_single_overlay(overlay, {})
                
                # Set initial visibility based on overlay.show property
                if artist and hasattr(artist, 'set_visible'):
                    artist.set_visible(overlay.show)
                    print(f"[RenderPlotOp] Set initial visibility for {overlay.id}: {overlay.show}")
                elif artist and isinstance(artist, list):
                    # Handle multiple artists (e.g., multiple lines)
                    for a in artist:
                        if hasattr(a, 'set_visible'):
                            a.set_visible(overlay.show)
                    print(f"[RenderPlotOp] Set initial visibility for {overlay.id} (multiple artists): {overlay.show}")
            
            self._refresh_plot_widget()
            print(f"[RenderPlotOp] Updated {len(overlays)} statistical overlays")
            return True
            
        except Exception as e:
            print(f"[RenderPlotOp] Error updating statistical overlays: {e}")
            return False
    
    def toggle_overlay_visibility(self, overlay_id: str, visible: bool):
        """Toggle the visibility of an overlay element."""
        if overlay_id in self.overlay_artists:
            artist = self.overlay_artists[overlay_id]
            
            # Handle single artist
            if hasattr(artist, 'set_visible'):
                artist.set_visible(visible)
                print(f"[RenderPlotOp] Set overlay {overlay_id} visibility to {visible}")
            # Handle multiple artists (e.g., multiple lines)
            elif isinstance(artist, list):
                for a in artist:
                    if hasattr(a, 'set_visible'):
                        a.set_visible(visible)
                print(f"[RenderPlotOp] Set overlay {overlay_id} (multiple artists) visibility to {visible}")
            
            # Refresh the plot to show changes
            self._refresh_plot_widget()
            return True
        else:
            print(f"[RenderPlotOp] Overlay {overlay_id} not found in artist registry")
            return False
    
    def _apply_performance_options(self, scatter_data: List[Dict[str, Any]], plot_config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Apply performance options to scatter data before rendering.
        
        NOTE: Performance options only apply to scatter plots. Other plot types
        (bar, stacked_area, histogram, line) ignore performance options.
        """
        try:
            processed_data = []
            
            print(f"[RenderPlotOp] Applying performance options to scatter plot: {plot_config}")
            
            for pair_data in scatter_data:
                # Make a copy to avoid modifying original data
                processed_pair = pair_data.copy()
                
                # Apply max points downsampling
                if plot_config.get('max_points_enabled', False):
                    max_points = plot_config.get('max_points', 5000)
                    processed_pair = self._downsample_data(processed_pair, max_points)
                
                # Store density mode for later use in rendering
                processed_pair['density_mode'] = plot_config.get('density_mode', 'scatter')
                processed_pair['bins'] = plot_config.get('bins', 50)
                
                processed_data.append(processed_pair)
            
            return processed_data
            
        except Exception as e:
            print(f"[RenderPlotOp] Error applying performance options: {e}")
            return scatter_data  # Return original data on error
    
    def _downsample_data(self, pair_data: Dict[str, Any], max_points: int) -> Dict[str, Any]:
        """Downsample data to maximum number of points using random sampling."""
        try:
            x_data = pair_data.get('x_data', [])
            y_data = pair_data.get('y_data', [])
            
            if len(x_data) <= max_points:
                return pair_data  # No downsampling needed
            
            print(f"[RenderPlotOp] Downsampling from {len(x_data)} to {max_points} points for pair {pair_data.get('pair_name', 'unknown')}")
            
            # Use random sampling to preserve distribution
            import numpy as np
            indices = np.random.choice(len(x_data), max_points, replace=False)
            indices = np.sort(indices)  # Keep chronological order
            
            # Apply downsampling
            if isinstance(x_data, list):
                pair_data['x_data'] = [x_data[i] for i in indices]
                pair_data['y_data'] = [y_data[i] for i in indices]
            else:
                # Handle numpy arrays
                pair_data['x_data'] = x_data[indices]
                pair_data['y_data'] = y_data[indices]
            
            # Update point count
            pair_data['n_points'] = max_points
            
            return pair_data
            
        except Exception as e:
            print(f"[RenderPlotOp] Error downsampling data: {e}")
            return pair_data  # Return original data on error
    
    def _render_scatter_points(self, pair_data: Dict[str, Any]):
        """Render traditional scatter points (existing logic)."""
        if not self.current_axes:
            print("[RenderPlotOp] No axes available for scatter points")
            return
            
        try:
            x_data = pair_data.get('x_data', [])
            y_data = pair_data.get('y_data', [])
            
            # Get styling from pair data
            color = pair_data.get('color', '#1f77b4')
            marker = pair_data.get('marker', 'o')
            alpha = pair_data.get('alpha', 0.6)
            pair_name = pair_data.get('pair_name', 'Unknown')
            pair_id = pair_data.get('pair_id', pair_name)
            n_points = pair_data.get('n_points', len(x_data))
            
            # Get additional styling properties from pair data
            marker_size = pair_data.get('marker_size', 50)
            edge_color = pair_data.get('edge_color', '#000000')
            edge_width = pair_data.get('edge_width', 1.0)
            z_order = pair_data.get('z_order', 0)
            
            # Create scatter plot and store artist reference
            artist = self.current_axes.scatter(
                x_data, y_data,
                c=color, marker=marker, alpha=alpha,
                s=marker_size, 
                edgecolors=edge_color if edge_width > 0 else 'none',
                linewidths=edge_width,
                zorder=z_order,
                label=f"{pair_name} (n={n_points})"
            )
            
            # Store artist reference for visibility toggling
            self.pair_artists[pair_id] = artist
            
            print(f"[RenderPlotOp] Rendered scatter points for pair '{pair_name}' with {n_points} points")
            print(f"[RenderPlotOp] Styling: size={marker_size}, alpha={alpha}, edge={edge_color}@{edge_width}")
            
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering scatter points: {e}")
            import traceback
            traceback.print_exc()
    
    def _render_hexbin_plot(self, pair_data: Dict[str, Any]):
        """Render hexbin density plot."""
        if not self.current_axes:
            print("[RenderPlotOp] No axes available for hexbin plot")
            return
            
        try:
            x_data = pair_data.get('x_data', [])
            y_data = pair_data.get('y_data', [])
            bins = pair_data.get('bins', 50)
            pair_name = pair_data.get('pair_name', 'Unknown')
            pair_id = pair_data.get('pair_id', pair_name)
            
            # Create hexbin plot
            artist = self.current_axes.hexbin(
                x_data, y_data, 
                gridsize=bins, 
                cmap='Blues',
                alpha=0.8,
                mincnt=1  # Minimum count to display
            )
            
            # Store artist reference for visibility toggling
            self.pair_artists[pair_id] = artist
            
            print(f"[RenderPlotOp] Rendered hexbin plot for pair '{pair_name}' with {bins} bins")
            
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering hexbin plot: {e}")
    
    def _render_kde_plot(self, pair_data: Dict[str, Any]):
        """Render KDE density plot."""
        try:
            x_data = pair_data.get('x_data', [])
            y_data = pair_data.get('y_data', [])
            pair_name = pair_data.get('pair_name', 'Unknown')
            pair_id = pair_data.get('pair_id', pair_name)
            
            # Try to use seaborn for KDE if available
            try:
                import seaborn as sns
                import numpy as np
                
                # Convert to numpy arrays if needed
                x_array = np.array(x_data)
                y_array = np.array(y_data)
                
                # Create KDE plot
                artist = sns.kdeplot(
                    x=x_array, y=y_array,
                    ax=self.current_axes,
                    fill=True,
                    alpha=0.6,
                    levels=10
                )
                
                # Store artist reference for visibility toggling
                self.pair_artists[pair_id] = artist
                
                print(f"[RenderPlotOp] Rendered KDE plot for pair '{pair_name}' using seaborn")
                
            except ImportError:
                # Fallback to matplotlib contour plot
                import numpy as np
                from scipy.stats import gaussian_kde
                
                # Convert to numpy arrays
                x_array = np.array(x_data)
                y_array = np.array(y_data)
                
                # Create KDE
                kde = gaussian_kde(np.vstack([x_array, y_array]))
                
                # Create grid for contour plot
                x_min, x_max = x_array.min(), x_array.max()
                y_min, y_max = y_array.min(), y_array.max()
                
                x_grid = np.linspace(x_min, x_max, 50)
                y_grid = np.linspace(y_min, y_max, 50)
                X, Y = np.meshgrid(x_grid, y_grid)
                
                # Evaluate KDE on grid
                positions = np.vstack([X.ravel(), Y.ravel()])
                Z = kde(positions).reshape(X.shape)
                
                # Create contour plot
                artist = self.current_axes.contourf(X, Y, Z, levels=10, alpha=0.6, cmap='Blues')
                
                # Store artist reference for visibility toggling
                self.pair_artists[pair_id] = artist
                
                print(f"[RenderPlotOp] Rendered KDE plot for pair '{pair_name}' using matplotlib/scipy")
                
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering KDE plot: {e}")
            # Fallback to scatter plot
            self._render_scatter_points(pair_data)
    
    def _render_bar_plot(self, scatter_data: List[Dict[str, Any]], plot_config: Dict[str, Any]):
        """Render bar chart plot."""
        try:
            print(f"[RenderPlotOp] Rendering bar plot with {len(scatter_data)} data series")
            
            # For bar plots, we expect data to be in categories and values format
            for i, pair_data in enumerate(scatter_data):
                x_data = pair_data.get('x_data', [])
                y_data = pair_data.get('y_data', [])
                
                if len(x_data) == 0 or len(y_data) == 0:
                    continue
                
                pair_name = pair_data.get('pair_name', 'Unknown')
                pair_id = pair_data.get('pair_id', pair_name)
                color = pair_data.get('color', '#1f77b4')
                alpha = pair_data.get('alpha', 0.8)
                
                # Create bar plot
                artist = self.current_axes.bar(
                    x_data, y_data,
                    color=color, alpha=alpha,
                    label=pair_name
                )
                
                # Store artist reference for visibility toggling
                self.pair_artists[pair_id] = artist
                
                print(f"[RenderPlotOp] Rendered bar plot for pair '{pair_name}' with {len(x_data)} bars")
                
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering bar plot: {e}")
    
    def _render_stacked_area_plot(self, scatter_data: List[Dict[str, Any]], plot_config: Dict[str, Any]):
        """Render stacked area chart."""
        try:
            print(f"[RenderPlotOp] Rendering stacked area plot with {len(scatter_data)} data series")
            
            if not scatter_data:
                return
            
            # Collect all x_data and y_data for stacking
            x_data = scatter_data[0].get('x_data', [])
            y_data_series = []
            labels = []
            colors = []
            
            for pair_data in scatter_data:
                y_data_series.append(pair_data.get('y_data', []))
                labels.append(pair_data.get('pair_name', 'Unknown'))
                colors.append(pair_data.get('color', '#1f77b4'))
            
            # Create stacked area plot
            artist = self.current_axes.stackplot(
                x_data, *y_data_series,
                labels=labels,
                colors=colors,
                alpha=0.7
            )
            
            # Store artist reference for visibility toggling
            for i, pair_data in enumerate(scatter_data):
                pair_id = pair_data.get('pair_id', pair_data.get('pair_name', 'Unknown'))
                self.pair_artists[pair_id] = artist[i] if i < len(artist) else artist
            
            print(f"[RenderPlotOp] Rendered stacked area plot with {len(y_data_series)} series")
            
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering stacked area plot: {e}")
    
    def _render_histogram_plot(self, scatter_data: List[Dict[str, Any]], plot_config: Dict[str, Any]):
        """Render histogram plot with advanced styling and binning support."""
        try:
            print(f"[RenderPlotOp] Rendering histogram plot with {len(scatter_data)} data series")
            
            for pair_data in scatter_data:
                x_data = pair_data.get('x_data', [])
                
                if len(x_data) == 0:
                    continue
                
                pair_name = pair_data.get('pair_name', 'Unknown')
                pair_id = pair_data.get('pair_id', pair_name)
                
                # Use histogram-specific styling from pair_analyzer.py
                fill_color = pair_data.get('bar_fill_color', pair_data.get('color', '#1f77b4'))
                edge_color = pair_data.get('bar_edge_color', '#000000')
                edge_width = pair_data.get('bar_edge_width', 0.8)
                alpha = pair_data.get('bar_alpha', pair_data.get('alpha', 0.7))
                z_order = pair_data.get('bar_z_order', pair_data.get('z_order', 0))
                
                # Get histogram-specific parameters
                bins = pair_data.get('bins', 30)
                bin_edges = pair_data.get('bin_edges', None)
                histogram_type = pair_data.get('histogram_type', 'counts')
                
                # Determine binning strategy
                if bin_edges is not None:
                    # Use pre-computed bin edges from comparison class
                    bins_param = bin_edges
                    print(f"[RenderPlotOp] Using pre-computed bin edges: {len(bin_edges)} edges")
                else:
                    # Use bin count
                    bins_param = bins
                    print(f"[RenderPlotOp] Using bin count: {bins}")
                
                # Determine normalization (density parameter)
                if histogram_type == 'density':
                    density = True
                    print(f"[RenderPlotOp] Using density normalization")
                elif histogram_type == 'probability':
                    # For probability, we'll normalize manually after hist() call
                    density = False
                    print(f"[RenderPlotOp] Using probability normalization")
                else:
                    # Default to counts
                    density = False
                    print(f"[RenderPlotOp] Using counts normalization")
                
                # Create histogram with advanced parameters
                n, bins_edges, patches = self.current_axes.hist(
                    x_data, 
                    bins=bins_param,
                    color=fill_color, 
                    alpha=alpha,
                    edgecolor=edge_color if edge_width > 0 else 'none',
                    linewidth=edge_width,
                    label=pair_name,
                    zorder=z_order,
                    density=density
                )
                
                # Handle probability normalization manually
                if histogram_type == 'probability':
                    # Convert counts to probabilities
                    total_count = sum(n)
                    if total_count > 0:
                        # Update patch heights to show probabilities
                        for patch, prob in zip(patches, n / total_count):
                            patch.set_height(prob)
                    print(f"[RenderPlotOp] Converted counts to probabilities (total: {total_count})")
                
                # Store artist reference for visibility toggling
                self.pair_artists[pair_id] = patches
                
                # Log detailed information
                print(f"[RenderPlotOp] Rendered histogram for pair '{pair_name}':")
                print(f"  - Bins: {len(bins_edges)-1 if hasattr(bins_edges, '__len__') else bins}")
                print(f"  - Type: {histogram_type}")
                print(f"  - Fill: {fill_color}, Edge: {edge_color}@{edge_width}")
                print(f"  - Alpha: {alpha}, Z-order: {z_order}")
                print(f"  - Data range: [{min(x_data):.3f}, {max(x_data):.3f}]")
                print(f"  - Sample count: {len(x_data)}")
                
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering histogram plot: {e}")
            import traceback
            traceback.print_exc()
    
    def _render_line_plot(self, scatter_data: List[Dict[str, Any]], plot_config: Dict[str, Any]):
        """Render line plot."""
        try:
            print(f"[RenderPlotOp] Rendering line plot with {len(scatter_data)} data series")
            
            for pair_data in scatter_data:
                x_data = pair_data.get('x_data', [])
                y_data = pair_data.get('y_data', [])
                
                if len(x_data) == 0 or len(y_data) == 0:
                    continue
                
                pair_name = pair_data.get('pair_name', 'Unknown')
                pair_id = pair_data.get('pair_id', pair_name)
                color = pair_data.get('color', '#1f77b4')
                alpha = pair_data.get('alpha', 0.8)
                line_style = pair_data.get('line_style', 'solid')
                
                # Create line plot
                artist = self.current_axes.plot(
                    x_data, y_data,
                    color=color, alpha=alpha,
                    linestyle=line_style,
                    linewidth=2,
                    label=pair_name
                )
                
                # Store artist reference for visibility toggling
                self.pair_artists[pair_id] = artist[0]  # plot returns a list
                
                print(f"[RenderPlotOp] Rendered line plot for pair '{pair_name}' with {len(x_data)} points")
                
        except Exception as e:
            print(f"[RenderPlotOp] Error rendering line plot: {e}")


class ComparisonWizardManager(QMainWindow):
    """
    Manager for the Comparison Wizard Window
    Handles the lifecycle and integration with the main application
    """
    
    # Signals for communication with main window
    comparison_completed = Signal(dict)  # Emitted when comparison is completed
    wizard_closed = Signal()  # Emitted when wizard is closed
    
    def __init__(self, file_manager=None, channel_manager=None, signal_bus=None, parent=None, selected_file_id=None):
        super().__init__(parent)
        
        # Store managers
        self.file_manager = file_manager
        self.channel_manager = channel_manager
        self.signal_bus = signal_bus
        self.selected_file_id = selected_file_id
        
        # Create pair selection operation with selected file priority
        self.pair_selection_op = PairSelectionOp(file_manager, channel_manager, selected_file_id)
        
        # Initialize pair manager
        from pair_manager import PairManager
        self.pair_manager = PairManager(max_pairs=100)
        
        # Initialize comparison registry
        self._initialize_comparison_registry()
        
        # Initialize PairAnalyzer
        self.pair_analyzer = PairAnalyzer(comparison_registry=ComparisonRegistry)
        
        # Initialize RenderPlotOp (will be connected to plot widget later)
        self.render_plot_op = RenderPlotOp()
        
        # Debouncing timer for analysis updates
        self.analysis_timer = QTimer()
        self.analysis_timer.setSingleShot(True)
        self.analysis_timer.timeout.connect(self._perform_analysis)
        self.analysis_debounce_delay = 300  # 300ms delay
        
        # Create and setup wizard window
        self.comparison_wizard = ComparisonWizardWindow(
            file_manager=file_manager,
            channel_manager=channel_manager,
            signal_bus=signal_bus,
            parent=self
        )
        
        # Set the manager reference so the wizard can access manager methods
        self.comparison_wizard.manager = self
        
        # Refresh the wizard with actual comparison methods from registry
        self._refresh_wizard_comparison_methods()
        
        # Set window properties
        self.setWindowTitle("Data Comparison Wizard")
        self.setMinimumSize(1200, 800)
        
        # Set the comparison wizard as the central widget
        self.setCentralWidget(self.comparison_wizard)
        
        # Connect signals from the comparison wizard
        self._connect_wizard_signals()
        
        # State tracking
        self.is_active = False
        self.aligned_pairs = {}
        
        # Initialize method configuration operation
        self.method_config_op = None
        
        print("[ComparisonWizardManager] Initialized with PairManager integration")
    
    def _initialize_comparison_registry(self):
        """Initialize the comparison registry and load all comparison methods"""
        try:
            # Load all comparison methods
            success = load_all_comparisons()
            if success:
                # Get all available comparison methods
                self.comparison_methods = ComparisonRegistry.all_comparisons()
                print(f"[ComparisonWizardManager] Loaded {len(self.comparison_methods)} comparison methods: {self.comparison_methods}")
            else:
                print("[ComparisonWizardManager] Failed to load comparison methods")
                self.comparison_methods = []
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error initializing comparison registry: {e}")
            self.comparison_methods = []
    
    def get_comparison_methods(self):
        """Get list of available comparison method names"""
        return self.comparison_methods
    
    def get_comparison_class(self, method_name):
        """Get a comparison class by method name"""
        try:
            return ComparisonRegistry.get(method_name)
        except Exception as e:
            print(f"[ComparisonWizardManager] Error getting comparison class for {method_name}: {e}")
            return None
    
    def get_comparison_info(self, method_name):
        """Get detailed info about a comparison method"""
        comparison_cls = self.get_comparison_class(method_name)
        if comparison_cls:
            return comparison_cls.get_info()
        return None
        
    def _connect_wizard_signals(self):
        """Connect signals from the comparison wizard"""
        # Connect pair management signals
        self.comparison_wizard.pair_added.connect(self._on_pair_added)
        self.comparison_wizard.pair_deleted.connect(self._on_pair_deleted)
        self.comparison_wizard.plot_generated.connect(self._on_plot_generated)
        
    def show(self):
        """Show the comparison wizard manager"""
        try:
            # Validate that we have data to work with
            if not self._validate_data_availability():
                return
                
            # Populate the wizard with available data
            self._populate_wizard_data()
            
            # Mark as active and show
            self.is_active = True
            super().show()
            
            # Emit signal that wizard is now active
            if self.signal_bus:
                self.signal_bus.emit('comparison_wizard_opened', {})
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open Comparison Wizard: {str(e)}")
            traceback.print_exc()
            
    def closeEvent(self, event):
        """Handle window close event"""
        try:
            # Mark as inactive
            self.is_active = False
            
            # Emit close signal
            self.wizard_closed.emit()
            
            # Emit signal bus event if available
            if self.signal_bus:
                self.signal_bus.emit('comparison_wizard_closed', {})
                
            # Accept the close event
            event.accept()
            
        except Exception as e:
            print(f"Error in comparison wizard close: {e}")
            event.accept()
            
    def _validate_data_availability(self):
        """Validate that we have sufficient data for comparison"""
        if not self.file_manager:
            QMessageBox.warning(self, "No File Manager", "File manager is not available.")
            return False
            
        if not self.channel_manager:
            QMessageBox.warning(self, "No Channel Manager", "Channel manager is not available.")
            return False
            
        # Check if we have any files loaded
        try:
            files = self.file_manager.get_all_files()
            if not files:
                QMessageBox.information(self, "No Data", "Please load some files first before using the comparison wizard.")
                return False
        except AttributeError:
            QMessageBox.warning(self, "Invalid File Manager", "File manager does not have required methods.")
            return False
            
        # Check if we have any channels available
        try:
            channels = self.channel_manager.get_all_channels()
            if not channels:
                QMessageBox.information(self, "No Channels", "No channels are available for comparison. Please ensure files are properly parsed.")
                return False
        except AttributeError:
            QMessageBox.warning(self, "Invalid Channel Manager", "Channel manager does not have required methods.")
            return False
            
        return True
        
    def _populate_wizard_data(self):
        """Populate the wizard with available file and channel data, then autopopulate intelligently"""
        try:
            # Get available files and channels
            if not self.file_manager or not self.channel_manager:
                return
                
            files = self.file_manager.get_all_files()
            channels = self.channel_manager.get_all_channels()
            
            # Update the wizard's file and channel dropdowns
            self._update_file_dropdowns(files)
            self._update_channel_dropdowns(channels)
            
            # The wizard now handles its own comparison method population in its constructor
            # No need to call _refresh_wizard_comparison_methods() here
            
            # Call intelligent autopopulation
            self._autopopulate_intelligent_channels()
            
            print(f"[ComparisonWizardManager] Populated wizard with {len(files)} files and {len(channels)} channels")
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error populating wizard data: {e}")
    
    def _refresh_wizard_comparison_methods(self):
        """Refresh the comparison methods in the wizard"""
        try:
            if hasattr(self.comparison_wizard, 'refresh_comparison_methods'):
                success = self.comparison_wizard.refresh_comparison_methods(self.comparison_methods)
                if success:
                    print(f"[ComparisonWizardManager] Successfully refreshed wizard with {len(self.comparison_methods)} methods")
                    # Ensure correlation method is selected by default
                    self._ensure_correlation_default_selection()
                else:
                    print("[ComparisonWizardManager] Failed to refresh wizard methods")
            else:
                print("[ComparisonWizardManager] Wizard does not have refresh_comparison_methods method")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error refreshing wizard methods: {e}")
    
    def _ensure_correlation_default_selection(self):
        """Ensure the correlation method is selected by default in the wizard"""
        try:
            if hasattr(self.comparison_wizard, 'method_list'):
                # Find and select correlation method
                for i in range(self.comparison_wizard.method_list.count()):
                    item_text = self.comparison_wizard.method_list.item(i).text()
                    if 'correlation' in item_text.lower():
                        self.comparison_wizard.method_list.setCurrentRow(i)
                        print(f"[ComparisonWizardManager] Ensured correlation method is selected: {item_text}")
                        return
                
                # If correlation not found, select first item
                if self.comparison_wizard.method_list.count() > 0:
                    self.comparison_wizard.method_list.setCurrentRow(0)
                    first_item = self.comparison_wizard.method_list.item(0).text()
                    print(f"[ComparisonWizardManager] Selected first method as fallback: {first_item}")
                    
        except Exception as e:
            print(f"[ComparisonWizardManager] Error ensuring correlation default selection: {e}")
            
    def _update_file_dropdowns(self, files):
        """Update file dropdowns in the wizard, storing file objects as data"""
        try:
            file_combo_boxes = [self.comparison_wizard.ref_file_combo, self.comparison_wizard.test_file_combo]
            
            for combo in file_combo_boxes:
                # Store current selection
                current_file = combo.currentData() if combo.currentIndex() >= 0 else None
                
                # Clear and repopulate
                combo.clear()
                for file_obj in files:
                    combo.addItem(file_obj.filename, file_obj)
                
                # Try to restore previous selection
                if current_file:
                    for i in range(combo.count()):
                        if combo.itemData(i) == current_file:
                            combo.setCurrentIndex(i)
                            break
                            
        except Exception as e:
            print(f"Error updating file dropdowns: {e}")
            
    def _update_channel_dropdowns(self, channels):
        """Initialize channel dropdowns as empty - they will be populated when files are selected"""
        try:
            channel_combo_boxes = [self.comparison_wizard.ref_channel_combo, self.comparison_wizard.test_channel_combo]
            
            for combo in channel_combo_boxes:
                # Clear the combo box - channels will be populated when files are selected
                combo.clear()
                combo.addItem("Select a file first...", None)
                combo.setCurrentIndex(0)
                
            print(f"[ComparisonWizardManager] Initialized empty channel dropdowns - will be populated when files are selected")
                            
        except Exception as e:
            print(f"Error initializing channel dropdowns: {e}")

    def _autopopulate_intelligent_channels(self):
        """Auto-populate channels using intelligent selection"""
        try:
            # Log selected file priority if available
            if self.selected_file_id and self.file_manager:
                selected_file = self.file_manager.get_file(self.selected_file_id)
                if selected_file:
                    print(f"[ComparisonWizard] Using selected file priority: {selected_file.filename}")
            
            # Find best channel pair using the pair selection operation
            ref_channel, test_channel = self.pair_selection_op.find_intelligent_channel_pairs()
            
            if ref_channel and test_channel:
                # Set files first
                self._select_file_for_channel(ref_channel, self.comparison_wizard.ref_file_combo)
                self._select_file_for_channel(test_channel, self.comparison_wizard.test_file_combo)
                
                # Trigger channel dropdown updates based on selected files
                self._trigger_channel_dropdown_updates()
                
                # Set channels (now that channel dropdowns are filtered)
                self._select_channel_in_combo(ref_channel, self.comparison_wizard.ref_channel_combo)
                self._select_channel_in_combo(test_channel, self.comparison_wizard.test_channel_combo)
                
                print(f"[ComparisonWizard] Auto-selected: {ref_channel.legend_label} vs {test_channel.legend_label}")
                print(f"[ComparisonWizard] From files: {ref_channel.file_id} vs {test_channel.file_id}")
                
                # Generate pair name after channels are selected
                self._update_pair_name(ref_channel, test_channel)
            else:
                print("[ComparisonWizard] No suitable channel pairs found for auto-selection")
                
        except Exception as e:
            print(f"[ComparisonWizard] Error in intelligent channel selection: {e}")

    def _update_pair_name(self, ref_channel=None, test_channel=None):
        """Update pair name based on selected channels"""
        try:
            if ref_channel is None:
                ref_channel = self.comparison_wizard.ref_channel_combo.currentData()
            if test_channel is None:
                test_channel = self.comparison_wizard.test_channel_combo.currentData()
            
            if ref_channel and test_channel:
                # Call the wizard's pair name generation method
                if hasattr(self.comparison_wizard, '_generate_pair_name'):
                    new_name = self.comparison_wizard._generate_pair_name(ref_channel, test_channel)
                    self.comparison_wizard.pair_name_input.setText(new_name)
                    print(f"[ComparisonWizard] Updated pair name: {new_name}")
        except Exception as e:
            print(f"[ComparisonWizard] Error updating pair name: {e}")

    def _select_file_for_channel(self, channel, file_combo):
        """Select the file that contains the given channel"""
        for i in range(file_combo.count()):
            file_info = file_combo.itemData(i)
            if file_info and file_info.file_id == channel.file_id:
                file_combo.setCurrentIndex(i)
                break

    def _select_channel_in_combo(self, channel, channel_combo):
        """Select the given channel in the combo box"""
        channel_name = getattr(channel, 'legend_label', None) or getattr(channel, 'channel_id', str(channel))
        for i in range(channel_combo.count()):
            if channel_combo.itemText(i) == channel_name:
                channel_combo.setCurrentIndex(i)
                break
    
    def _trigger_channel_dropdown_updates(self):
        """Trigger channel dropdown updates based on currently selected files"""
        try:
            # Update ref channel dropdown based on ref file selection
            ref_file = self.comparison_wizard.ref_file_combo.currentData()
            if ref_file and hasattr(self.comparison_wizard, '_update_channel_dropdown_for_file'):
                self.comparison_wizard._update_channel_dropdown_for_file(ref_file, self.comparison_wizard.ref_channel_combo)
                print(f"[ComparisonWizardManager] Triggered ref channel dropdown update for file: {ref_file.filename}")
            
            # Update test channel dropdown based on test file selection
            test_file = self.comparison_wizard.test_file_combo.currentData()
            if test_file and hasattr(self.comparison_wizard, '_update_channel_dropdown_for_file'):
                self.comparison_wizard._update_channel_dropdown_for_file(test_file, self.comparison_wizard.test_channel_combo)
                print(f"[ComparisonWizardManager] Triggered test channel dropdown update for file: {test_file.filename}")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error triggering channel dropdown updates: {e}")
            
    def _on_pair_added(self, pair_data):
        """Handle when a comparison pair is added"""
        try:
            # Log the addition
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"Comparison pair added: {pair_data.get('name', 'Unnamed')}")
            
            # Get channel information for early duplicate check
            ref_channel = pair_data.get('ref_channel')
            test_channel = pair_data.get('test_channel')
            
            if not ref_channel or not test_channel:
                error_msg = "Missing channel information for pair"
                print(f"[ComparisonWizardManager] ERROR: {error_msg}")
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append(f"❌ {error_msg}")
                return
            
            # Early duplicate detection - check BEFORE expensive alignment
            is_duplicate, existing_pair_name = self.pair_manager.is_duplicate_pair(
                ref_channel.channel_id, test_channel.channel_id,
                ref_channel.file_id, test_channel.file_id
            )
            
            if is_duplicate:
                error_msg = f"Duplicate pair detected: '{pair_data.get('name', 'Unnamed')}' conflicts with existing pair '{existing_pair_name}'. Use the delete icon to remove the existing pair first."
                print(f"[ComparisonWizardManager] BLOCKED: {error_msg}")
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append(f"❌ {error_msg}")
                
                # Emit blocked signal
                self.comparison_completed.emit({
                    'type': 'pair_add_blocked',
                    'data': pair_data,
                    'error': error_msg
                })
                return
            
            # Proceed with alignment since no duplicate was found
            print(f"[ComparisonWizardManager] No duplicate found, proceeding with alignment for: {pair_data.get('name')}")
            
            # Check if alignment was successful
            alignment_result = pair_data.get('alignment_result')
            if alignment_result and alignment_result.success:
                print(f"[ComparisonWizardManager] Successfully aligned pair: {pair_data.get('name')}")
                print(f"[ComparisonWizardManager] Aligned data shape: {alignment_result.ref_data.shape}")
                
                # Log quality information if available
                if hasattr(alignment_result, 'quality_info'):
                    quality = alignment_result.quality_info
                    print(f"[ComparisonWizardManager] Data quality - Length: {quality.get('data_length', 'N/A')}, "
                          f"Correlation: {quality.get('correlation', 'N/A'):.3f}")
                
                # Create Pair object with aligned data
                from pair import Pair
                
                pair = Pair(
                    name=pair_data.get('name'),
                    ref_channel_id=ref_channel.channel_id if ref_channel else None,
                    test_channel_id=test_channel.channel_id if test_channel else None,
                    ref_file_id=ref_channel.file_id if ref_channel else None,
                    test_file_id=test_channel.file_id if test_channel else None,
                    ref_channel_name=getattr(ref_channel, 'legend_label', None) if ref_channel else None,
                    test_channel_name=getattr(test_channel, 'legend_label', None) if test_channel else None,
                    alignment_config=pair_data.get('alignment_config'),
                    description=pair_data.get('description', ''),
                    metadata=pair_data.get('metadata', {})
                )
                
                # Set aligned data (this will automatically compute basic stats)
                pair.set_aligned_data(
                    alignment_result.ref_data,
                    alignment_result.test_data
                )
                
                # Log computed statistics
                if pair.r_squared is not None:
                    print(f"[ComparisonWizardManager] Pair statistics - R²: {pair.r_squared:.3f}, "
                          f"Correlation: {pair.correlation:.3f}, "
                          f"Mean diff: {pair.mean_difference:.3f}")
                    
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(
                            f"📊 Statistics: R²={pair.r_squared:.3f}, r={pair.correlation:.3f}"
                        )
                
                # Add pair to PairManager (duplicate check is redundant now since we checked early)
                success, message = self.pair_manager.add_pair(pair)
                if success:
                    # Store the Pair object for later use
                    if not hasattr(self, 'aligned_pairs'):
                        self.aligned_pairs = {}
                    self.aligned_pairs[pair_data.get('name')] = {
                        'pair_object': pair,
                        'ref_channel': ref_channel,
                        'test_channel': test_channel,
                        'alignment_result': alignment_result,
                        'alignment_params': pair_data.get('alignment_params')
                    }
                    
                    # Update channels table (immediate UI update)
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(f"📋 Channels table updated")
                    
                    # Trigger debounced analysis
                    self._trigger_analysis_update()
                    
                    # Log capacity information
                    capacity_info = self.pair_manager.get_capacity_info()
                    if capacity_info['at_warning_level']:
                        print(f"[ComparisonWizardManager] WARNING: {capacity_info['current_count']}/{capacity_info['max_pairs']} pairs in manager")
                    
                    # Log success message
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(f"✅ {message}")
                    
                    # Emit signal to main window AFTER successful addition
                    self.comparison_completed.emit({
                        'type': 'pair_added',
                        'data': pair_data
                    })
                    
                else:
                    # Pair addition was blocked (likely duplicate)
                    print(f"[ComparisonWizardManager] Pair addition blocked: {message}")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(f"❌ {message}")
                    
                    # Emit blocked signal
                    self.comparison_completed.emit({
                        'type': 'pair_add_blocked',
                        'data': pair_data,
                        'error': message
                    })
            else:
                error_msg = "Unknown alignment error"
                if alignment_result and alignment_result.error_message:
                    error_msg = alignment_result.error_message
                print(f"[ComparisonWizardManager] Alignment failed for pair: {pair_data.get('name')} - {error_msg}")
                
                # Log error message
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append(f"❌ Alignment failed: {error_msg}")
                
                # Emit error signal
                self.comparison_completed.emit({
                    'type': 'pair_add_failed',
                    'data': pair_data,
                    'error': error_msg
                })
            
        except Exception as e:
            print(f"Error handling pair addition: {e}")
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"❌ Error adding pair: {e}")
            
            # Emit error signal
            self.comparison_completed.emit({
                'type': 'pair_add_failed',
                'data': pair_data if 'pair_data' in locals() else {},
                'error': str(e)
            })
            
    def _on_pair_deleted(self, pair_data):
        """Handle when a comparison pair is deleted"""
        try:
            # Extract pair name from the signal data
            pair_name = pair_data.get('pair_name', 'Unknown') if isinstance(pair_data, dict) else str(pair_data)
            
            print(f"[ComparisonWizardManager] Pair deletion signal received for: {pair_name}")
            
            # Find and remove the pair from PairManager
            if hasattr(self, 'pair_manager') and self.pair_manager:
                # Find the pair by name
                pair_to_remove = None
                for pair in self.pair_manager.get_all_pairs():
                    if getattr(pair, 'name', '') == pair_name:
                        pair_to_remove = pair
                        break
                
                if pair_to_remove:
                    # Remove from PairManager
                    success = self.pair_manager.remove_pair(pair_to_remove.pair_id)
                    if success:
                        print(f"[ComparisonWizardManager] Successfully removed pair '{pair_name}' from PairManager")
                        
                        # Remove from aligned_pairs if it exists there
                        if hasattr(self, 'aligned_pairs') and pair_name in self.aligned_pairs:
                            del self.aligned_pairs[pair_name]
                            print(f"[ComparisonWizardManager] Removed pair '{pair_name}' from aligned_pairs")
                        
                        # Trigger analysis update to recompute plot with remaining pairs
                        self._trigger_analysis_update()
                        
                        # Log the deletion
                        if hasattr(self.comparison_wizard, 'info_output'):
                            self.comparison_wizard.info_output.append(f"🗑️ Pair '{pair_name}' deleted - plot updated")
                    else:
                        print(f"[ComparisonWizardManager] Failed to remove pair '{pair_name}' from PairManager")
                        if hasattr(self.comparison_wizard, 'info_output'):
                            self.comparison_wizard.info_output.append(f"⚠️ Failed to delete pair '{pair_name}' from backend")
                else:
                    print(f"[ComparisonWizardManager] Could not find pair '{pair_name}' in PairManager")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(f"⚠️ Pair '{pair_name}' not found in backend")
            else:
                print(f"[ComparisonWizardManager] No PairManager available for deletion")
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append(f"⚠️ No PairManager available for deletion")
                
            # Emit signal to main window
            self.comparison_completed.emit({
                'type': 'pair_deleted',
                'pair_name': pair_name
            })
            
        except Exception as e:
            print(f"Error handling pair deletion: {e}")
            import traceback
            traceback.print_exc()
            
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"❌ Error deleting pair: {e}")
            
    def _on_plot_generated(self, plot_data):
        """Handle when a plot is generated"""
        try:
            # Log the plot generation
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append("📊 Plot generated successfully")
                
            # Emit signal to main window
            self.comparison_completed.emit({
                'type': 'plot_generated',
                'data': plot_data
            })
            
        except Exception as e:
            print(f"Error handling plot generation: {e}")
            
    def get_active_comparisons(self):
        """Get currently active comparisons"""
        try:
            # Check if we have aligned_pairs stored in manager
            if hasattr(self, 'aligned_pairs'):
                return list(self.aligned_pairs.keys())
            
            # Fallback: check if wizard has any pairs in its table
            if hasattr(self.comparison_wizard, 'channels_table'):
                active_pairs = []
                for i in range(self.comparison_wizard.channels_table.rowCount()):
                    name_item = self.comparison_wizard.channels_table.item(i, 2)  # Pair name column
                    if name_item:
                        active_pairs.append(name_item.text())
                return active_pairs
            
            return []
        except Exception as e:
            print(f"Error getting active comparisons: {e}")
            return []
            
    def refresh_data(self):
        """Refresh the wizard with current data"""
        try:
            if self.is_active:
                self._populate_wizard_data()
        except Exception as e:
            print(f"Error refreshing data: {e}")

    def perform_alignment(self, ref_channel, test_channel, alignment_params):
        """Perform alignment using enhanced DataAligner with comprehensive validation and fallback"""
        try:
            print(f"[DEBUG] perform_alignment: Starting alignment")
            print(f"[DEBUG] perform_alignment: ref_channel type={type(ref_channel)}")
            print(f"[DEBUG] perform_alignment: test_channel type={type(test_channel)}")
            print(f"[DEBUG] perform_alignment: alignment_params={alignment_params}")
            
            from data_aligner import DataAligner
            
            # Create data aligner instance
            aligner = DataAligner()
            
            # Perform alignment with enhanced validation and fallback
            alignment_result = aligner.align_from_wizard_params(
                ref_channel, test_channel, alignment_params
            )
            
            # Get enhanced alignment statistics
            alignment_stats = aligner.get_alignment_stats()
            
            # Log enhanced statistics and feedback
            if alignment_result.success:
                print(f"[ComparisonWizardManager] Alignment completed successfully")
                print(f"[ComparisonWizardManager] Aligned data length: {len(alignment_result.ref_data)}")
                print(f"[ComparisonWizardManager] Alignment stats: {alignment_stats}")
                
                # Log datetime conversion if it happened
                if alignment_stats.get('datetime_conversions', 0) > 0:
                    print(f"[ComparisonWizardManager] Datetime conversion applied")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append("🕐 Applied datetime conversion to alignment")
                
                # Log fallback usage if it happened
                if alignment_stats.get('fallback_usage', 0) > 0:
                    print(f"[ComparisonWizardManager] Fallback strategies used")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append("⚠️ Used fallback alignment strategies")
                
                # Log quality metrics if available
                if hasattr(alignment_result, 'quality_metrics') and alignment_result.quality_metrics:
                    quality = alignment_result.quality_metrics
                    print(f"[ComparisonWizardManager] Quality metrics: {quality}")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        retention_ref = quality.get('ref_data_retention', 0) * 100
                        retention_test = quality.get('test_data_retention', 0) * 100
                        self.comparison_wizard.info_output.append(f"📊 Data retention: ref={retention_ref:.1f}%, test={retention_test:.1f}%")
                
                # Log warnings if available
                if hasattr(alignment_result, 'warnings') and alignment_result.warnings:
                    for warning in alignment_result.warnings:
                        print(f"[ComparisonWizardManager] Alignment warning: {warning}")
                        if hasattr(self.comparison_wizard, 'info_output'):
                            self.comparison_wizard.info_output.append(f"⚠️ {warning}")
            else:
                # Enhanced error reporting
                error_msg = alignment_result.error_message or "Unknown alignment error"
                print(f"[ComparisonWizardManager] Alignment failed: {error_msg}")
                
                # Log fallback attempts
                if alignment_stats.get('fallback_usage', 0) > 0:
                    print(f"[ComparisonWizardManager] Fallback strategies attempted but failed")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append("❌ Fallback strategies attempted but failed")
                
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append(f"❌ Alignment failed: {error_msg}")
            
            return alignment_result
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Critical error in perform_alignment: {e}")
            # Return error result
            from data_aligner import AlignmentResult
            return AlignmentResult(
                ref_data=np.array([]),
                test_data=np.array([]),
                success=False,
                error_message=f"Critical alignment failure: {e}"
            )

    def _validate_alignment_result(self, alignment_result, ref_channel, test_channel):
        """DEPRECATED: DataAligner now has comprehensive validation built-in"""
        # DataAligner now handles all validation internally with fallback strategies
        # This method is kept for backward compatibility but not used
        return {
            'valid': True,
            'quality_info': {'data_length': len(alignment_result.ref_data) if alignment_result.ref_data is not None else 0}
        }

    def _calculate_data_quality(self, ref_data, test_data):
        """Calculate quality metrics for aligned data"""
        try:
            ref_mean = np.mean(ref_data)
            ref_std = np.std(ref_data)
            test_mean = np.mean(test_data)
            test_std = np.std(test_data)
            
            # Calculate correlation coefficient
            if ref_std > 0 and test_std > 0:
                correlation = np.corrcoef(ref_data, test_data)[0, 1]
                if np.isnan(correlation):
                    correlation = 0.0
            else:
                correlation = 0.0
            
            # Calculate range ratios
            ref_range = np.max(ref_data) - np.min(ref_data)
            test_range = np.max(test_data) - np.min(test_data)
            
            return {
                'ref_mean': ref_mean,
                'ref_std': ref_std,
                'ref_range': ref_range,
                'test_mean': test_mean,
                'test_std': test_std,
                'test_range': test_range,
                'correlation': correlation,
                'data_length': len(ref_data)
            }
            
        except Exception as e:
            print(f"Error calculating data quality: {e}")
            return {
                'ref_mean': 0.0,
                'ref_std': 0.0,
                'ref_range': 0.0,
                'test_mean': 0.0,
                'test_std': 0.0,
                'test_range': 0.0,
                'correlation': 0.0,
                'data_length': 0
            }

    def suggest_alignment_parameters(self, mode=None):
        """Public method to suggest alignment parameters - can be called from wizard window"""
        try:
            ref_channel = self.comparison_wizard.ref_channel_combo.currentData()
            test_channel = self.comparison_wizard.test_channel_combo.currentData()
            if not ref_channel or not test_channel:
                return None
            
            op = DataAlignmentOp(ref_channel, test_channel)
            return op.suggest_alignment(mode=mode)
        except Exception as e:
            print(f"Error suggesting alignment parameters: {e}")
            return None

    def update_alignment_section(self):
        """Update Data Alignment section based on current channel selection and mode"""
        try:
            ref_channel = self.comparison_wizard.ref_channel_combo.currentData()
            test_channel = self.comparison_wizard.test_channel_combo.currentData()
            if not ref_channel or not test_channel:
                return
            mode = self.comparison_wizard.alignment_mode_combo.currentText()
            op = DataAlignmentOp(ref_channel, test_channel)
            params = op.suggest_alignment(mode=mode)
            # Update the Data Alignment UI fields
            alignment_mode_combo = self.comparison_wizard.alignment_mode_combo
            index_group = self.comparison_wizard.index_group
            time_group = self.comparison_wizard.time_group
            # Set mode
            alignment_mode_combo.setCurrentText(params['mode'])
            if params['mode'] == 'Index-Based':
                index_group.show()
                time_group.hide()
                self.comparison_wizard.start_index_spin.setValue(params.get('start_index', 0))
                self.comparison_wizard.end_index_spin.setValue(params.get('end_index', 0))
                self.comparison_wizard.index_offset_spin.setValue(params.get('offset', 0))
            else:
                index_group.hide()
                time_group.show()
                self.comparison_wizard.start_time_spin.setValue(params.get('start_time', 0.0))
                self.comparison_wizard.end_time_spin.setValue(params.get('end_time', 0.0))
                self.comparison_wizard.time_offset_spin.setValue(params.get('offset', 0.0))
                interp = params.get('interpolation', 'nearest')
                interp_index = self.comparison_wizard.interpolation_combo.findText(interp)
                if interp_index >= 0:
                    self.comparison_wizard.interpolation_combo.setCurrentIndex(interp_index)
                self.comparison_wizard.resolution_spin.setValue(params.get('resolution', 0.1))
        except Exception as e:
            print(f"Error updating alignment section: {e}")

    def get_pair_manager_stats(self):
        """Get statistics from the PairManager"""
        return self.pair_manager.get_stats()
    
    def get_pair_manager_capacity(self):
        """Get capacity information from the PairManager"""
        return self.pair_manager.get_capacity_info()
    
    def get_data_aligner_stats(self):
        """Get DataAligner statistics for debugging/monitoring"""
        try:
            from data_aligner import DataAligner
            aligner = DataAligner()
            return aligner.get_alignment_stats()
        except Exception as e:
            print(f"Error getting DataAligner stats: {e}")
            return {}
    
    def clear_data_aligner_cache(self):
        """Clear DataAligner cache - useful for debugging"""
        try:
            from data_aligner import DataAligner
            aligner = DataAligner()
            aligner.clear_cache()
            print("[ComparisonWizardManager] DataAligner cache cleared")
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append("🧹 DataAligner cache cleared")
        except Exception as e:
            print(f"Error clearing DataAligner cache: {e}")
    
    def validate_alignment_compatibility(self, ref_channel, test_channel):
        """Pre-validate if channels are compatible for alignment"""
        try:
            from data_aligner import DataAligner
            aligner = DataAligner()
            
            # Use the enhanced DataAligner validation
            ref_validation = aligner.data_validator.validate_channel_data(ref_channel)
            test_validation = aligner.data_validator.validate_channel_data(test_channel)
            
            compatibility_info = {
                'ref_valid': ref_validation.is_valid,
                'test_valid': test_validation.is_valid,
                'ref_issues': ref_validation.issues,
                'test_issues': test_validation.issues,
                'ref_warnings': ref_validation.warnings,
                'test_warnings': test_validation.warnings,
                'ref_quality_score': ref_validation.data_quality_score,
                'test_quality_score': test_validation.data_quality_score,
                'overall_compatible': ref_validation.is_valid and test_validation.is_valid
            }
            
            return compatibility_info
            
        except Exception as e:
            print(f"Error validating alignment compatibility: {e}")
            return {
                'ref_valid': False,
                'test_valid': False,
                'ref_issues': [f"Validation error: {e}"],
                'test_issues': [f"Validation error: {e}"],
                'ref_warnings': [],
                'test_warnings': [],
                'ref_quality_score': 0.0,
                'test_quality_score': 0.0,
                'overall_compatible': False
            }
    
    def _trigger_analysis_update(self):
        """Trigger debounced analysis update"""
        print("[ComparisonWizardManager] Triggering debounced analysis update...")
        self.analysis_timer.stop()  # Cancel any pending analysis
        self.analysis_timer.start(self.analysis_debounce_delay)
    
    def _perform_analysis(self):
        """Perform analysis using PairAnalyzer and update plot/overlays"""
        try:
            print("[ComparisonWizardManager] Computing analysis...")
            
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append("🔄 Computing analysis...")
            
            # Capture current method configuration
            method_config = self._capture_method_config()
            if not method_config:
                print("[ComparisonWizardManager] No method configuration available")
                return
            
            # Clear cache to ensure parameter changes are reflected
            if hasattr(self.pair_analyzer, 'clear_cache'):
                self.pair_analyzer.clear_cache()
                print("[ComparisonWizardManager] Cleared PairAnalyzer cache for refresh")
            
            # Run PairAnalyzer
            analysis_results = self.pair_analyzer.analyze(self.pair_manager, method_config)
            
            # Store analysis results for export
            self._last_analysis_results = analysis_results
            
            # DEBUG: Check what we got from PairAnalyzer
            print(f"[ComparisonWizardManager] DEBUG: PairAnalyzer returned analysis_results keys: {list(analysis_results.keys())}")
            scatter_data = analysis_results.get('scatter_data', [])
            overlays = analysis_results.get('overlays', [])
            print(f"[ComparisonWizardManager] DEBUG: scatter_data length: {len(scatter_data)}")
            print(f"[ComparisonWizardManager] DEBUG: overlays length: {len(overlays)}")
            if scatter_data:
                print(f"[ComparisonWizardManager] DEBUG: First scatter_data item keys: {list(scatter_data[0].keys())}")
                print(f"[ComparisonWizardManager] DEBUG: First scatter_data x_data length: {len(scatter_data[0].get('x_data', []))}")
                print(f"[ComparisonWizardManager] DEBUG: First scatter_data y_data length: {len(scatter_data[0].get('y_data', []))}")
            
            # Check for errors
            errors = analysis_results.get('errors', {})
            if errors:
                for pair_id, error_msg in errors.items():
                    print(f"[ComparisonWizardManager] Error in pair {pair_id}: {error_msg}")
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(f"❌ Error in pair {pair_id}: {error_msg}")
            
            # Update plot area using RenderPlotOp
            plot_widget = self._get_plot_widget()
            if plot_widget:
                self.render_plot_op.plot_widget = plot_widget
                
                # Get performance options from method config
                performance_options = getattr(method_config, 'performance_options', {})
                
                # Use preserved overlays in the analysis results
                analysis_results_with_preserved = analysis_results.copy()
                preserved_overlays = self._preserve_custom_text_across_analysis(analysis_results.get('overlays', []))
                analysis_results_with_preserved['overlays'] = preserved_overlays
                
                success = self.render_plot_op.render(analysis_results_with_preserved, plot_config=performance_options)
                if success:
                    print("[ComparisonWizardManager] Plot rendered successfully")
                else:
                    print("[ComparisonWizardManager] Plot rendering failed")
            
            # Update overlay table
            overlays = analysis_results.get('overlays', [])
            
            # Preserve custom text from existing overlays
            preserved_overlays = self._preserve_custom_text_across_analysis(overlays)
            
            self._update_overlay_table(preserved_overlays)
            
            # Store overlays for visibility toggle handling
            self._last_overlays = preserved_overlays
            
            # Log completion
            n_pairs = analysis_results.get('n_pairs_processed', 0)
            cache_stats = analysis_results.get('cache_stats', {})
            
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"✅ Analysis complete. {n_pairs} pairs processed. Cache: {cache_stats.get('hits', 0)} hits, {cache_stats.get('misses', 0)} misses")
            
            print(f"[ComparisonWizardManager] Analysis complete: {n_pairs} pairs, cache stats: {cache_stats}")
            
            # Emit signal to indicate analysis refresh is complete
            self.comparison_completed.emit({
                'type': 'analysis_refreshed',
                'n_pairs': n_pairs,
                'cache_stats': cache_stats
            })
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error in analysis: {e}")
            import traceback
            traceback.print_exc()
            
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"❌ Analysis error: {e}")
            
            # Emit error signal
            self.comparison_completed.emit({
                'type': 'analysis_refresh_failed',
                'error': str(e)
            })
    
    def _trigger_hybrid_visibility_update(self, pair_id: str, visible: bool):
        """Trigger hybrid visibility update: immediate visual + recalculated stats."""
        print(f"[ComparisonWizardManager] Hybrid visibility update for pair {pair_id}: {visible}")
        
        try:
            # Step 1: Immediate visual toggle
            if hasattr(self, 'render_plot_op'):
                success = self.render_plot_op.toggle_pair_visibility(pair_id, visible)
                if not success:
                    print(f"[ComparisonWizardManager] Visual toggle failed, falling back to full analysis")
                    self._perform_analysis()
                    return
            
            # Step 2: Recalculate stats from visible pairs only using cached results
            self._recalculate_stats_for_visible_pairs()
            
            # Step 3: Update overlay table
            overlays = getattr(self, '_last_overlays', [])
            if overlays:
                self._update_overlay_table(overlays)
            
            print(f"[ComparisonWizardManager] Hybrid update completed for pair {pair_id}")
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error in hybrid update, falling back to full analysis: {e}")
            self._perform_analysis()
    
    def _update_statistical_overlays_from_cache(self):
        """Update statistical overlays using cached PairAnalyzer results."""
        try:
            # Get cached results from PairAnalyzer
            if not hasattr(self, 'pair_analyzer') or not self.pair_analyzer:
                print("[ComparisonWizardManager] No PairAnalyzer available for cache update")
                return
            
            # Get current method config
            method_config = self._capture_method_config()
            if not method_config:
                print("[ComparisonWizardManager] No method config available")
                return
            
            # Get visible pairs from PairManager
            visible_pairs = []
            if hasattr(self, 'pair_manager') and self.pair_manager:
                visible_pairs = self.pair_manager.get_visible_pairs()
            
            if not visible_pairs:
                print("[ComparisonWizardManager] No visible pairs for stats update")
                return
            
            # Get comparison class
            comparison_cls = self.pair_analyzer._get_comparison_class(method_config.method_name)
            if not comparison_cls:
                print("[ComparisonWizardManager] No comparison class found")
                return
            
            # Collect cached stats results for visible pairs only
            all_stats_results = []
            for pair in visible_pairs:
                cache_key = self.pair_analyzer._generate_pair_cache_key(pair.pair_id, method_config)
                cached_result = self.pair_analyzer._get_cached_result(cache_key)
                if cached_result:
                    all_stats_results.append(cached_result['stats_results'])
            
            if not all_stats_results:
                print("[ComparisonWizardManager] No cached stats results available")
                return
            
            # Generate new overlays from filtered stats
            overlays = self.pair_analyzer._generate_overlays(comparison_cls, all_stats_results, method_config)
            
            # Update overlays in renderer
            if hasattr(self, 'render_plot_op'):
                self.render_plot_op.update_statistical_overlays(overlays)
                self._last_overlays = overlays  # Store for overlay table update
            
            print(f"[ComparisonWizardManager] Updated {len(overlays)} statistical overlays from cache")
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error updating statistical overlays from cache: {e}")
            import traceback
            traceback.print_exc()
    
    def _recalculate_stats_for_visible_pairs(self):
        """Recalculate statistics and overlays from visible pairs only using cached results."""
        try:
            print("[ComparisonWizardManager] Recalculating stats for visible pairs...")
            
            # Get current method configuration
            method_config = self._capture_method_config()
            if not method_config:
                print("[ComparisonWizardManager] No method configuration for recalculation")
                return
            
            # Use PairAnalyzer to recombine visible pairs
            analysis_results = self.pair_analyzer.recombine_visible_pairs(self.pair_manager, method_config)
            
            if not analysis_results:
                print("[ComparisonWizardManager] No results from recombination")
                return
            
            # Update overlays using RenderPlotOp
            overlays = analysis_results.get('overlays', [])
            if hasattr(self, 'render_plot_op') and self.render_plot_op:
                success = self.render_plot_op.update_statistical_overlays(overlays)
                if success:
                    print(f"[ComparisonWizardManager] Updated {len(overlays)} overlays from visible pairs")
                    
                    # Update overlay table
                    self._update_overlay_table(overlays)
                    self._last_overlays = overlays
                    
                    # Log to console
                    n_pairs = analysis_results.get('n_pairs_processed', 0)
                    if hasattr(self.comparison_wizard, 'info_output'):
                        self.comparison_wizard.info_output.append(f"🔄 Updated stats from {n_pairs} visible pairs")
                else:
                    print("[ComparisonWizardManager] Failed to update overlays from visible pairs")
            else:
                print("[ComparisonWizardManager] No render plot operation available for overlay update")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error recalculating stats for visible pairs: {e}")
            import traceback
            traceback.print_exc()
    
    def _on_overlay_visibility_changed(self, overlay_id: str, state: int):
        """Handle overlay visibility checkbox state change"""
        try:
            print(f"[ComparisonWizardManager] Overlay visibility changed: {overlay_id}, state: {state}")
            
            # Convert Qt checkbox state to boolean
            is_visible = state == Qt.CheckState.Checked.value
            
            # Update overlay object property
            overlays = getattr(self, '_last_overlays', [])
            overlay_updated = False
            
            for overlay in overlays:
                if overlay.id == overlay_id:
                    overlay.show = is_visible
                    overlay_updated = True
                    print(f"[ComparisonWizardManager] Updated overlay {overlay_id} show property to {is_visible}")
                    break
            
            if not overlay_updated:
                print(f"[ComparisonWizardManager] Warning: Could not find overlay {overlay_id} to update")
            
            # Toggle overlay visibility in renderer
            if hasattr(self, 'render_plot_op'):
                success = self.render_plot_op.toggle_overlay_visibility(overlay_id, is_visible)
                if success:
                    print(f"[ComparisonWizardManager] Successfully toggled overlay {overlay_id} visibility")
                else:
                    print(f"[ComparisonWizardManager] Failed to toggle overlay {overlay_id} visibility")
            
            # Update info output
            if hasattr(self.comparison_wizard, 'info_output'):
                visibility_text = "shown" if is_visible else "hidden"
                self.comparison_wizard.info_output.append(f"👁️ Overlay {overlay_id} {visibility_text}")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error handling overlay visibility change: {e}")
            import traceback
            traceback.print_exc()
    
    def _capture_method_config(self) -> Optional[PairAnalyzerMethodConfig]:
        """Capture current method configuration from wizard"""
        try:
            if not self.comparison_wizard:
                return None
            
            # Get current method name
            method_name = self.comparison_wizard.get_current_method_name()
            if not method_name:
                print("[ComparisonWizardManager] No method selected")
                return None
            
            # Get current parameters
            parameters = self.comparison_wizard.get_current_parameters()
            
            # Get performance options from wizard
            performance_options = {}
            if hasattr(self.comparison_wizard, 'get_performance_options'):
                performance_options = self.comparison_wizard.get_performance_options()
                print(f"[ComparisonWizardManager] Captured performance options: {performance_options}")
            
            # Get scripts from wizard - only capture if modified
            plot_script = None
            stats_script = None
            
            try:
                if hasattr(self.comparison_wizard, 'plot_script_text') and hasattr(self.comparison_wizard, 'script_tracker'):
                    current_plot_script = self.comparison_wizard.plot_script_text.toPlainText()
                    if current_plot_script and current_plot_script.strip():
                        # Check if script has been modified using ScriptChangeTracker
                        if self.comparison_wizard.script_tracker.is_plot_script_modified(current_plot_script):
                            plot_script = current_plot_script
                            print(f"[ComparisonWizardManager] Captured MODIFIED plot script ({len(plot_script)} chars)")
                        else:
                            print(f"[ComparisonWizardManager] Plot script present but not modified - using default")
                        
                if hasattr(self.comparison_wizard, 'stats_script_text') and hasattr(self.comparison_wizard, 'script_tracker'):
                    current_stats_script = self.comparison_wizard.stats_script_text.toPlainText()
                    if current_stats_script and current_stats_script.strip():
                        # Check if script has been modified using ScriptChangeTracker
                        if self.comparison_wizard.script_tracker.is_stats_script_modified(current_stats_script):
                            stats_script = current_stats_script
                            print(f"[ComparisonWizardManager] Captured MODIFIED stats script ({len(stats_script)} chars)")
                        else:
                            print(f"[ComparisonWizardManager] Stats script present but not modified - using default")
                        
            except Exception as e:
                print(f"[ComparisonWizardManager] Error capturing scripts: {e}")
                plot_script = None
                stats_script = None
            
            # Create method config
            method_config = PairAnalyzerMethodConfig(
                method_name=method_name,
                parameters=parameters,
                plot_script=plot_script,
                stats_script=stats_script,
                performance_options=performance_options
            )
            
            print(f"[ComparisonWizardManager] Captured method config: {method_name} with {len(parameters)} parameters")
            print(f"[ComparisonWizardManager] Modified plot_script: {'Yes' if plot_script else 'No'}")
            print(f"[ComparisonWizardManager] Modified stats_script: {'Yes' if stats_script else 'No'}")
            
            return method_config
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error capturing method config: {e}")
            return None
    
    def _get_plot_widget(self):
        """Get the plot widget from the wizard for rendering"""
        try:
            if hasattr(self.comparison_wizard, 'canvas'):
                print("[ComparisonWizardManager] Found plot canvas in wizard")
                return self.comparison_wizard.canvas
            else:
                print("[ComparisonWizardManager] No plot canvas found in wizard")
                return None
        except Exception as e:
            print(f"[ComparisonWizardManager] Error getting plot widget: {e}")
            return None
    
    def _update_overlay_table(self, overlays: List[Overlay]):
        """Update the overlay table in the wizard with new overlays"""
        try:
            print(f"[ComparisonWizardManager] Updating overlay table with {len(overlays)} overlays - v2")
            
            # Get the overlay table from the wizard
            if not hasattr(self.comparison_wizard, 'overlay_table'):
                print("[ComparisonWizardManager] No overlay table found in wizard")
                return
            
            overlay_table = self.comparison_wizard.overlay_table
            
            # Clear existing rows
            overlay_table.setRowCount(0)
            
            # Add each overlay to the table
            for i, overlay in enumerate(overlays):
                overlay_table.insertRow(i)
                
                # Show checkbox
                checkbox = QCheckBox()
                checkbox.setChecked(overlay.show)
                checkbox.stateChanged.connect(lambda state, overlay_id=overlay.id: self._on_overlay_visibility_changed(overlay_id, state))
                overlay_table.setCellWidget(i, 0, checkbox)
                
                # Style preview based on overlay properties
                style_widget = self._create_overlay_style_preview(overlay)
                overlay_table.setCellWidget(i, 1, style_widget)
                
                # Name
                name_item = QTableWidgetItem(overlay.name)
                name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
                overlay_table.setItem(i, 2, name_item)
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                
                paint_btn = QPushButton("🎨")
                paint_btn.setMaximumSize(24, 24)
                paint_btn.setToolTip("Edit overlay style")
                paint_btn.clicked.connect(self._create_paint_handler(overlay))
                actions_layout.addWidget(paint_btn)
                
                info_btn = QPushButton("ℹ️")
                info_btn.setMaximumSize(24, 24)
                info_btn.clicked.connect(self._create_info_handler(overlay))
                actions_layout.addWidget(info_btn)
                
                overlay_table.setCellWidget(i, 3, actions_widget)
                
                print(f"[ComparisonWizardManager] Overlay {i}: {overlay.name} (ID: {overlay.id}, type: {overlay.type}) - Show: {overlay.show}")
            
            # Store the overlays for reference
            self._last_overlays = overlays
            
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"📊 Overlay table updated with {len(overlays)} overlays")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error updating overlay table: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_overlay_style_preview(self, overlay: Overlay) -> QWidget:
        """Create a style preview widget for an overlay based on its properties."""
        try:
            # Create the preview widget
            widget = QLabel()
            widget.setFixedSize(60, 20)
            widget.setAlignment(Qt.AlignCenter)
            
            # Get overlay properties
            overlay_type = overlay.type
            style = overlay.style or {}
            
            # Create pixmap for drawing
            pixmap = QPixmap(60, 20)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing)
            
            # Get color from style or use default
            color = style.get('color', '#1f77b4')  # Default blue
            alpha = style.get('alpha', 0.8)
            
            # Set up color with alpha
            qcolor = QColor(color)
            qcolor.setAlphaF(alpha)
            
            # Draw preview based on overlay type
            if overlay_type == 'line':
                # Draw a horizontal line
                pen = QPen(qcolor, style.get('linewidth', 2))
                pen.setStyle(self._get_qt_line_style(style.get('linestyle', '-')))
                painter.setPen(pen)
                painter.drawLine(10, 10, 50, 10)
                
            elif overlay_type == 'hline':
                # Draw a horizontal line (same as line type)
                pen = QPen(qcolor, style.get('linewidth', 2))
                pen.setStyle(self._get_qt_line_style(style.get('linestyle', '-')))
                painter.setPen(pen)
                painter.drawLine(10, 10, 50, 10)
                
            elif overlay_type == 'vline':
                # Draw a vertical line
                pen = QPen(qcolor, style.get('linewidth', 2))
                pen.setStyle(self._get_qt_line_style(style.get('linestyle', '-')))
                painter.setPen(pen)
                painter.drawLine(30, 5, 30, 15)
                
            elif overlay_type == 'text':
                # Draw a text icon
                pen = QPen(qcolor, 1)
                painter.setPen(pen)
                painter.setFont(widget.font())
                painter.drawText(pixmap.rect(), Qt.AlignCenter, "T")
                
            elif overlay_type == 'fill':
                # Draw a filled rectangle
                brush = QBrush(qcolor)
                painter.setBrush(brush)
                painter.setPen(QPen(qcolor, 1))
                painter.drawRect(15, 5, 30, 10)
                
            elif overlay_type == 'marker':
                # Draw a marker symbol
                marker = style.get('marker', 'o')
                self._draw_marker(painter, marker, qcolor, 30, 10, 6)
                
            else:
                # Default: show type abbreviation
                pen = QPen(qcolor, 1)
                painter.setPen(pen)
                painter.setFont(widget.font())
                type_abbr = overlay_type[:2].upper() if overlay_type else "?"
                painter.drawText(pixmap.rect(), Qt.AlignCenter, type_abbr)
            
            painter.end()
            widget.setPixmap(pixmap)
            
            # Add tooltip with style information
            tooltip = self._create_style_tooltip(overlay)
            widget.setToolTip(tooltip)
            
            return widget
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error creating style preview: {e}")
            # Fallback to simple label
            fallback = QLabel("—")
            fallback.setStyleSheet("color: gray; font-size: 12px;")
            fallback.setAlignment(Qt.AlignCenter)
            fallback.setToolTip(f"Error creating preview: {e}")
            return fallback
    
    def _get_qt_line_style(self, linestyle: str) -> Qt.PenStyle:
        """Convert matplotlib line style to Qt pen style."""
        style_map = {
            '-': Qt.SolidLine,
            '--': Qt.DashLine,
            ':': Qt.DotLine,
            '-.': Qt.DashDotLine
        }
        return style_map.get(linestyle, Qt.SolidLine)
    
    def _draw_marker(self, painter: QPainter, marker: str, color: QColor, x: int, y: int, size: int):
        """Draw a marker symbol."""
        pen = QPen(color, 1)
        brush = QBrush(color)
        painter.setPen(pen)
        painter.setBrush(brush)
        
        if marker == 'o':  # Circle
            painter.drawEllipse(x - size, y - size, size * 2, size * 2)
        elif marker == 's':  # Square
            painter.drawRect(x - size, y - size, size * 2, size * 2)
        elif marker == '^':  # Triangle up
            points = [QPoint(x, y - size), QPoint(x - size, y + size), QPoint(x + size, y + size)]
            painter.drawPolygon(points)
        elif marker == 'D':  # Diamond
            points = [QPoint(x, y - size), QPoint(x + size, y), QPoint(x, y + size), QPoint(x - size, y)]
            painter.drawPolygon(points)
        else:  # Default to circle
            painter.drawEllipse(x - size, y - size, size * 2, size * 2)
    
    def _create_style_tooltip(self, overlay: Overlay) -> str:
        """Create a tooltip showing overlay style information."""
        try:
            style = overlay.style or {}
            lines = [f"Type: {overlay.type}"]
            
            # Add style properties
            if 'color' in style:
                lines.append(f"Color: {style['color']}")
            if 'alpha' in style:
                lines.append(f"Alpha: {style['alpha']:.2f}")
            if 'linewidth' in style:
                lines.append(f"Line Width: {style['linewidth']}")
            if 'linestyle' in style:
                lines.append(f"Line Style: {style['linestyle']}")
            if 'fontsize' in style:
                lines.append(f"Font Size: {style['fontsize']}")
            
            # Add functional properties
            if 'label' in style:
                lines.append(f"Label: {style['label']}")
            if 'position' in style:
                lines.append(f"Position: {style['position']}")
            
            return "\n".join(lines)
            
        except Exception as e:
            return f"Error creating tooltip: {e}"
    
    def _on_overlay_paint_clicked(self, overlay: Overlay):
        """Handle paint button click for overlay styling"""
        try:
            print(f"[ComparisonWizardManager] Opening overlay wizard for: {overlay.name} (type: {overlay.type}) - ID: {overlay.id}")
            print(f"[ComparisonWizardManager] Current overlay style: {overlay.style}")
            print(f"[ComparisonWizardManager] Overlay data: {overlay.data}")
            
            # OverlayWizard is already imported at the top
            
            # Check if overlay type is supported
            supported_types = ['line', 'text', 'fill', 'hline', 'vline']
            if overlay.type not in supported_types:
                print(f"[ComparisonWizardManager] Overlay type '{overlay.type}' not supported by wizard")
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append(f"⚠️ Overlay type '{overlay.type}' not supported for styling")
                return
            
            # Store the overlay reference for the signal handler
            self._current_editing_overlay = overlay
            
            # Open the overlay wizard with the correct overlay type
            wizard_type = overlay.type
            print(f"[ComparisonWizardManager] Opening wizard for overlay type: {wizard_type}")
            wizard = OverlayWizard(wizard_type, overlay.style, self)
            
            # For text overlays, prefill the text content
            if wizard_type == 'text':
                # Get the current text content from overlay
                current_text = overlay.display_text
                if current_text:
                    wizard.set_initial_text(current_text)
                    print(f"[ComparisonWizardManager] Prefilled text overlay with {len(current_text)} characters")
                else:
                    print(f"[ComparisonWizardManager] No text content found for overlay {overlay.id}")
            
            wizard.style_updated.connect(self._on_overlay_style_updated)
            result = wizard.exec()
            
            if result == QDialog.Accepted:
                print(f"[ComparisonWizardManager] Overlay wizard accepted for: {overlay.name}")
                print(f"[ComparisonWizardManager] New style from wizard: {wizard.overlay_style}")
                # Final update and cleanup
                self._update_overlay_style(overlay, wizard.overlay_style)
                self._refresh_overlay_table()
            else:
                print(f"[ComparisonWizardManager] Overlay wizard cancelled for: {overlay.name}")
            
            # Clear the stored overlay reference
            self._current_editing_overlay = None
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error opening overlay wizard: {e}")
            import traceback
            traceback.print_exc()
    
    def _on_overlay_style_updated(self, new_style: dict):
        """Handle style update signal from overlay wizard"""
        try:
            print(f"[ComparisonWizardManager] Overlay style updated")
            print(f"[ComparisonWizardManager] New style received: {new_style}")
            
            # Use the stored overlay reference if available
            overlay = getattr(self, '_current_editing_overlay', None)
            
            if overlay:
                print(f"[ComparisonWizardManager] Found overlay reference: {overlay.name} (ID: {overlay.id})")
                print(f"[ComparisonWizardManager] Overlay type: {overlay.type}")
                
                # Handle custom text content for text overlays
                if overlay.type == 'text' and 'text_content' in new_style:
                    custom_text = new_style['text_content']
                    overlay.set_custom_text(custom_text)
                    print(f"[ComparisonWizardManager] Set custom text: {len(custom_text)} characters")
                    
                    # Remove text_content from style since it's handled separately
                    new_style = new_style.copy()
                    del new_style['text_content']
                
                # Update the overlay style
                self._update_overlay_style(overlay, new_style)
                # Refresh the table to show updated styling
                self._refresh_overlay_table()
                print(f"[ComparisonWizardManager] Applied style updates to overlay: {overlay.name}")
            else:
                print(f"[ComparisonWizardManager] No overlay reference available for style update")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error handling overlay style update: {e}")
            import traceback
            traceback.print_exc()
    
    def _update_overlay_style(self, overlay: Overlay, new_style: dict):
        """Update an overlay's style and refresh the plot"""
        try:
            print(f"[ComparisonWizardManager] _update_overlay_style called for: {overlay.name} (ID: {overlay.id})")
            print(f"[ComparisonWizardManager] Old style: {overlay.style}")
            print(f"[ComparisonWizardManager] New style: {new_style}")
            
            # Update the overlay's style
            overlay.style.update(new_style)
            print(f"[ComparisonWizardManager] Updated overlay style: {overlay.style}")
            
            # Update the artist in the renderer if it exists
            if hasattr(self, 'render_plot_op') and self.render_plot_op:
                print(f"[ComparisonWizardManager] Updating artist for overlay: {overlay.id}")
                # Remove the old artist
                if overlay.id in self.render_plot_op.overlay_artists:
                    old_artist = self.render_plot_op.overlay_artists[overlay.id]
                    if hasattr(old_artist, 'remove'):
                        old_artist.remove()
                    elif isinstance(old_artist, list):
                        for a in old_artist:
                            if hasattr(a, 'remove'):
                                a.remove()
                    del self.render_plot_op.overlay_artists[overlay.id]
                    print(f"[ComparisonWizardManager] Removed old artist for: {overlay.id}")
                
                # Re-render the overlay with new style
                new_artist = self.render_plot_op._render_single_overlay(overlay, {})
                if new_artist:
                    self.render_plot_op.overlay_artists[overlay.id] = new_artist
                    # Set visibility
                    if hasattr(new_artist, 'set_visible'):
                        new_artist.set_visible(overlay.show)
                    elif isinstance(new_artist, list):
                        for a in new_artist:
                            if hasattr(a, 'set_visible'):
                                a.set_visible(overlay.show)
                    print(f"[ComparisonWizardManager] Created new artist for: {overlay.id}")
                else:
                    print(f"[ComparisonWizardManager] Failed to create new artist for: {overlay.id}")
                
                # Refresh the plot widget to show the changes
                self.render_plot_op._refresh_plot_widget()
                print(f"[ComparisonWizardManager] Refreshed plot widget")
            else:
                print(f"[ComparisonWizardManager] No render_plot_op available")
            
            print(f"[ComparisonWizardManager] Updated overlay {overlay.name} with new style")
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error updating overlay style: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_paint_handler(self, overlay: Overlay):
        """Create a paint button handler for a specific overlay"""
        def handler():
            self._on_overlay_paint_clicked(overlay)
        return handler
    
    def _create_info_handler(self, overlay: Overlay):
        """Create an info button handler for a specific overlay"""
        def handler():
            self._show_overlay_info(overlay)
        return handler
    
    def _show_overlay_info(self, overlay: Overlay):
        """Show information about an overlay"""
        try:
            print(f"[ComparisonWizardManager] Showing info for overlay: {overlay.name}")
            # TODO: Implement overlay info dialog
        except Exception as e:
            print(f"[ComparisonWizardManager] Error showing overlay info: {e}")
    
    def _refresh_overlay_table(self):
        """Refresh the overlay table to show updated styling"""
        try:
            overlays = getattr(self, '_last_overlays', [])
            if overlays:
                self._update_overlay_table(overlays)
                print(f"[ComparisonWizardManager] Refreshed overlay table")
        except Exception as e:
            print(f"[ComparisonWizardManager] Error refreshing overlay table: {e}")
    
    def _preserve_custom_text_across_analysis(self, new_overlays):
        """Preserve custom text from existing overlays across analysis updates"""
        try:
            # Get existing overlays from the last analysis
            existing_overlays = getattr(self, '_last_overlays', [])
            
            # Create a mapping of existing overlays by ID
            existing_overlay_map = {overlay.id: overlay for overlay in existing_overlays}
            
            preserved_overlays = []
            
            for new_overlay in new_overlays:
                # Check if there's an existing overlay with the same ID
                if new_overlay.id in existing_overlay_map:
                    existing_overlay = existing_overlay_map[new_overlay.id]
                    
                    # For text overlays, preserve custom text state
                    if (new_overlay.type == 'text' and existing_overlay.type == 'text' and 
                        existing_overlay.is_custom_text):
                        
                        # Transfer custom text state to new overlay
                        new_overlay.set_custom_text(existing_overlay._custom_text)
                        print(f"[ComparisonWizardManager] Preserved custom text for overlay {new_overlay.id}")
                        
                        # Also preserve other text-related state
                        if hasattr(existing_overlay, '_original_text'):
                            new_overlay._original_text = existing_overlay._original_text
                
                preserved_overlays.append(new_overlay)
            
            print(f"[ComparisonWizardManager] Preserved custom text for {len(preserved_overlays)} overlays")
            return preserved_overlays
            
        except Exception as e:
            print(f"[ComparisonWizardManager] Error preserving custom text: {e}")
            import traceback
            traceback.print_exc()
            return new_overlays  # Return new overlays if preservation fails
    
    def _export_analysis_data(self):
        """Export current analysis data to Excel format"""
        try:
            # Check if analysis results exist
            if not hasattr(self, '_last_analysis_results') or not self._last_analysis_results:
                if hasattr(self.comparison_wizard, 'info_output'):
                    self.comparison_wizard.info_output.append("Export Data: Exports currently plotted data and statistics to Excel format")
                    self.comparison_wizard.info_output.append("Run analysis first to generate data for export")
                return
            
            # Get file save location
            from PySide6.QtWidgets import QFileDialog
            import datetime
            
            # Generate default filename with timestamp and method
            method_name = self._get_current_method_name()
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"comparison_data_{method_name}_{timestamp}.xlsx"
            
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Export Analysis Data",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not filename:
                return  # User cancelled
            
            # Export the data
            self._write_excel_file(filename)
            
            # Success message
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"Data exported successfully to: {filename}")
                
        except Exception as e:
            print(f"[ComparisonWizardManager] Error exporting analysis data: {e}")
            if hasattr(self.comparison_wizard, 'info_output'):
                self.comparison_wizard.info_output.append(f"Error exporting data: {e}")
    
    def _get_current_method_name(self):
        """Get the current comparison method name"""
        try:
            method_config = self._capture_method_config()
            if method_config:
                return method_config.method_name
            return "unknown"
        except:
            return "unknown"
    
    def _write_excel_file(self, filename):
        """Write analysis data to Excel file"""
        import pandas as pd
        import openpyxl
        from openpyxl import Workbook
        
        # Get current analysis results
        analysis_results = self._last_analysis_results
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Plot Data
        self._create_plot_data_sheet(wb, analysis_results)
        
        # Sheet 2: Statistics
        self._create_statistics_sheet(wb, analysis_results)
        
        # Sheet 3: Visible Pairs
        self._create_pairs_sheet(wb)
        
        # Sheet 4: Analysis Config
        self._create_config_sheet(wb)
        
        # Save the workbook
        wb.save(filename)
        print(f"[ComparisonWizardManager] Excel file saved: {filename}")
    
    def _create_plot_data_sheet(self, wb, analysis_results):
        """Create the plot data sheet"""
        ws = wb.create_sheet("Plot Data")
        
        # Headers
        ws.append(["Pair_Name", "Pair_ID", "X_Data", "Y_Data", "Point_Index"])
        
        # Add data from scatter_data
        scatter_data = analysis_results.get('scatter_data', [])
        for scatter_item in scatter_data:
            pair_name = scatter_item.get('pair_name', 'Unknown')
            pair_id = scatter_item.get('pair_id', 'Unknown')
            x_data = scatter_item.get('x_data', [])
            y_data = scatter_item.get('y_data', [])
            
            # Write each data point as a row
            for i, (x, y) in enumerate(zip(x_data, y_data)):
                ws.append([pair_name, pair_id, x, y, i])
    
    def _create_statistics_sheet(self, wb, analysis_results):
        """Create the statistics sheet"""
        ws = wb.create_sheet("Statistics")
        
        # Headers
        ws.append(["Statistic_Name", "Value", "Units"])
        
        # Add data from statistics
        statistics = analysis_results.get('statistics', {})
        for stat_name, stat_value in statistics.items():
            if stat_name != 'error':
                ws.append([stat_name, stat_value, ""])
    
    def _create_pairs_sheet(self, wb):
        """Create the visible pairs sheet"""
        ws = wb.create_sheet("Visible Pairs")
        
        # Headers
        ws.append(["Pair_Name", "Pair_ID", "Ref_Channel", "Test_Channel", "Ref_File", "Test_File", "Show", "Description"])
        
        # Add data from visible pairs
        if hasattr(self, 'pair_manager') and self.pair_manager:
            visible_pairs = self.pair_manager.get_visible_pairs()
            for pair in visible_pairs:
                ws.append([
                    pair.name,
                    pair.pair_id,
                    pair.ref_channel_name or "Unknown",
                    pair.test_channel_name or "Unknown", 
                    pair.ref_file_id or "Unknown",
                    pair.test_file_id or "Unknown",
                    pair.show,
                    pair.description or ""
                ])
    
    def _create_config_sheet(self, wb):
        """Create the analysis config sheet"""
        ws = wb.create_sheet("Analysis Config")
        
        # Headers
        ws.append(["Parameter", "Value"])
        
        # Add configuration data
        method_config = self._capture_method_config()
        if method_config:
            ws.append(["Method", method_config.method_name])
            ws.append(["Parameters", str(method_config.parameters)])
            
        # Add timestamp
        import datetime
        ws.append(["Export_Timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        # Add pair count
        if hasattr(self, 'pair_manager') and self.pair_manager:
            ws.append(["Visible_Pairs_Count", len(self.pair_manager.get_visible_pairs())])
            ws.append(["Total_Pairs_Count", len(self.pair_manager.get_all_pairs())])